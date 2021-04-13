import xml.etree.ElementTree as ET
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows

class Autor():
	def __init__(self, nome, periodo, qualis_file):
		super(Autor, self).__init__()
		self.nome = nome
		self.periodo = periodo
		self.qualis_file = qualis_file
		
		self.info = {"Ano":[], "Título":[], "Nome de Publicação":[], "ISSN":[], "Qualis":[]}
		self.autores_dic = {'Título':[]}
		self.sufixos = ['Jr.', 'Jr', 'Filho', 'Neto']
		
		self.myroot = self.get_XML_file()
		self.fill_info()

	def get_XML_file(self):
		mytree = ET.parse(f"Curriculos/{self.nome}.xml")
		return mytree.getroot()

	def add_authors(self, pub):
		autores = []
		for i in pub:
			if 'NOME-PARA-CITACAO' in i.attrib.keys():
				autor = i.attrib['NOME-PARA-CITACAO'] 
				autor = autor.title()
				if "," in autor:
					autor_array = autor.split(", ")
					primeiro_nome = autor_array[1].split(" ")[0]

					sobrenome = autor_array[0].split(" ")[-1]
					if sobrenome in self.sufixos:
						if len(autor_array[0].split(" ")) > 1:
							sobrenome = autor_array[0].split(" ")[-2] + " " + sobrenome
						else:
							sobrenome = autor_array[1].split(" ")[-1] + " " + sobrenome
				else:
					autor_array = autor.split(" ")
					primeiro_nome = autor_array[0]
					sobrenome = autor_array[-1]
					if sobrenome in self.sufixos:
						sobrenome = autor_array[-2] + " " + sobrenome

				autor = primeiro_nome + " "
				autor += sobrenome

				autores.append(autor)
		
		for pos2, autor in enumerate(autores):
			if f"{pos2+1}º Autor" not in self.autores_dic:
				self.autores_dic[f"{pos2+1}º Autor"] = []
				for i in range(len(self.autores_dic["Título"])-1):
					self.autores_dic[f"{pos2+1}º Autor"].append("-")

		autor_keys = []
		for key in self.autores_dic.keys():
			if "º Autor" in key:
				autor_keys.append(key)
		for pos3, key in enumerate(autor_keys):
			if pos3 <= (len(autores)-1):
				self.autores_dic[key].append(autores[pos3])
			else:
				self.autores_dic[key].append("-")


	def add_data(self, pub, ano):
		self.autores_dic['Título'].append(pub[0].attrib['TITULO-DO-ARTIGO'])

		self.info["Ano"].append(ano)
					
		self.info["Título"].append(pub[0].attrib['TITULO-DO-ARTIGO'])
		self.info["Nome de Publicação"].append(pub[1].attrib['TITULO-DO-PERIODICO-OU-REVISTA'])
		
		issn = pub[1].attrib['ISSN']
		self.info["ISSN"].append(issn)

		qualis = self.qualis_file.loc[self.qualis_file['ISSN'] == issn]
		try:
			qualis = qualis.reset_index()
			qualis = qualis["Estrato"][0]
		except:
			qualis = "-"

		self.info["Qualis"].append(qualis)

	def fill_info(self):

		publications = self.myroot[1][1] # Artigos publicados

		for pub in publications:
			ano = int(pub[0].attrib['ANO-DO-ARTIGO'])
			if ano >= 2017:
				# print(self.nome, periodo)
				if periodo[str(ano)[2:]] == True:
					
					self.add_data(pub, ano)

					self.add_authors(pub)


class ExcelFile(Workbook):
	def __init__(self, relatorios, autores_dic):
		super(ExcelFile, self).__init__()
		self.relatorios = relatorios
		self.autores_dic = autores_dic
		self.add_info()
		self.aplica_estilo()
		self.converte_valores()
		self.aplica_dimensoes()


	def add_info(self):
		ws = self.active # Primeiro sheet
		ws.title =  'Dicionário de Autores'
		for row in dataframe_to_rows(self.autores_dic, index=False, header=True): # Adiciona o dataframe ao sheet
			  ws.append(row)

		for pos, autor in enumerate(self.relatorios["Autor"]):
			ws = self.create_sheet(autor)
			for row in dataframe_to_rows(self.relatorios["Relatorio"][pos], index=False, header=True): # Adiciona o dataframe ao sheet
			    ws.append(row)

	def aplica_estilo(self):
		for ws in self.worksheets:
			# Estilo dos nomes das colunas
			for cell in ws[1]:
				cell.font = Font(bold=True) # Negrito
				cell.alignment = Alignment(horizontal='center', vertical='center') # Centralizado

			if ws.title != 'Dicionário de Autores':
				# Centraliza todas as colunas menos a de título e de publicação
				for col in ws.columns:
					if col[0].column_letter not in 'BC':
						for cell in col:
							cell.alignment = Alignment(horizontal='center', vertical='center')
			else:
				for col in ws.columns:
					if col[0].column_letter not in 'A':
						for cell in col:
							cell.alignment = Alignment(horizontal='center', vertical='center')

	def converte_valores(self):
		for ws in self.worksheets:
			if ws.title != 'Dicionário de Autores':
				for cell in ws['D']: # Converte colunas de ISSN
					try:
						cell.value = int(cell.value)
					except:
						pass

	def aplica_dimensoes(self):
		for ws in self.worksheets:
			if ws.title != 'Dicionário de Autores':
				# Define tamanho das colunas
				ws.column_dimensions['A'].width = 15
				ws.column_dimensions['B'].width = 35
				ws.column_dimensions['C'].width = 35
				ws.column_dimensions['D'].width = 15
				ws.column_dimensions['E'].width = 15
		
				ws.row_dimensions[1].height = 25 # Altura da linha dos nomes das colunas
			else:
				dic = {1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J', 11: 'K', 12: 'L', 13 : 'M', 14: 'N', 15: 'O', 16: 'P', 17: 'Q', 18: 'R'}
				ws.column_dimensions['A'].width = 35
				ws.column_dimensions['B'].width = 25
				i = 3
				while i <= ws.max_column:
					ws.column_dimensions[dic[i]].width = 25
					i += 1
				
				ws.row_dimensions[1].height = 25 # Altura da linha dos nomes das colunas

	def salva_arquivo(self):
		self.save('lattes_qualis.xlsx') # Salva o arquivo



if __name__ == '__main__':
	professores = pd.read_csv("Relatório Lattes - Docentes 2017-2020.CSV", sep=";", encoding='iso-8859-1')
	qualis_file = pd.read_csv("QualisCC_2013_2016.csv", sep=";", encoding='iso-8859-1')

	for i in range(len(qualis_file["ISSN"])):
		qualis_file["ISSN"][i] = qualis_file["ISSN"][i].replace("-", "")


	relatorios = {'Autor':[], 'Relatorio':[]}
	autores_dic = {}
	lista_autores_dic = []
	for pos, professor in enumerate(professores["Nome"]):
		periodo = {"17": True, "18": True, "19": True, "20": True}
		for key in periodo.keys():
			if str(professores[key][pos]) == "nan":
				periodo[key] = False

		autor = Autor(professor, periodo, qualis_file)
		relatorios['Autor'].append(professor)
		relatorios['Relatorio'].append(pd.DataFrame(autor.info))

		lista_autores_dic.append(autor.autores_dic)

	maior = lista_autores_dic[0]
	for dic in lista_autores_dic:
		if len(dic) > len(maior):
			maior = dic

	autores_dic = {}
	for key in maior.keys():
		autores_dic[key] = []
		for dic in lista_autores_dic:
			if key in dic:
				for value in dic[key]:
					autores_dic[key].append(value)
			else:
				for i in range(len(dic['Título'])):
					autores_dic[key].append("-")

	excel = ExcelFile(relatorios, pd.DataFrame(autores_dic))
	excel.salva_arquivo()

	
	