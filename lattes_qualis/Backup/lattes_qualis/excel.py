from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from verifica_autores import em_lista_autores, trata_exceçoes
from grafico import Graficos, Graficos_Anais_Periodicos
import pandas as pd
import re
import os
from valores import ND
from PyscopusModified import ScopusModified

class ExcelFile(Workbook):
	def __init__(self, relatorios, autores, todos, artppg, medias, indicadores, indicadores_geral, lista_egressos, lista_alunos, excecoes, per_a1_a4, per_a1_a4_ae, periodicos, anais, periodicos_metricas, anais_metricas):
		super(ExcelFile, self).__init__()
		self.relatorios = relatorios
		self.autores = autores
		self.todos = todos
		self.artppg = artppg
		self.medias = medias
		self.indicadores = indicadores
		self.indicadores_geral = indicadores_geral
		self.lista_egressos = lista_egressos
		self.lista_alunos = lista_alunos
		self.excecoes = excecoes
		self.per_a1_a4 = per_a1_a4
		self.per_a1_a4_ae = per_a1_a4_ae
		self.periodicos = periodicos
		self.anais = anais
		self.periodicos_metricas = periodicos_metricas
		self.anais_metricas = anais_metricas

		for pos, egresso in enumerate(self.lista_egressos):
			self.lista_egressos[pos].name = trata_exceçoes(egresso.name.strip())
		for pos, aluno in enumerate(self.lista_alunos):
			self.lista_alunos[pos].name = trata_exceçoes(aluno.name.strip())

		self.add_info()
		# self.altera_autores()
		self.aplica_estilo()
		# self.converte_valores()
		self.aplica_dimensoes()
		self.aplica_cores()
		self.aplica_filtros()


	def styled_cells(self, data, ws, pinta=True):
		for c in data:
			c = Cell(ws, column="A", row=1, value=c)
			if c.value != None and str(c.value) != "nan":
				if c.value == "Porcentagem alunos/egressos":
					c.value = "Porcentagem"
				if data[0] in ["Periódicos", "A1-A4", "A1", "A2", "A3", "A4", "Irestrito"]:
					c.font = Font(color='FFFAFA')
					c.fill = PatternFill(fill_type='solid', start_color='00B050', end_color='00B050')
				elif data[0] != "Outros" and data[0] != "Número médio de docentes" and pinta == True:
					c.fill = PatternFill(fill_type='solid', start_color='FFFFCC', end_color='FFFFCC')
				if c.value in ["Tipo/Qualis", "Quantidade", "Porcentagem", "Quantidade com alunos/egressos", "Índice", "Acumulado", "Média por docente", "Número médio de docentes", "Nome Periódico", "Qualis 2019/ISSN Impresso", "Qualis 2019/ISSN Online", "Métrica", "Qtd.", "Qtd. %"]:
					c.font = Font(bold=True)
				if c.value != "Número médio de docentes":
					bd = Side(border_style="thin", color="000000")
					c.border = Border(left=bd, top=bd, right=bd, bottom=bd)
					c.alignment = Alignment(horizontal='center', vertical='center')
			yield c

	def add_info(self):
		ws = self.active # Primeiro sheet
		ws.title = 'Parâmetros'
		ws.append(["Estrato", "Peso"])
		estratos = ["A1", "A2", "A3", "A4", "B1", "B2", "B3", "B4", ]
		pesos = [1.000, 0.875, 0.750, 0.625, 0.500, 0.200, 0.100, 0.050, ]
		for pos, estrato in enumerate(estratos):
			ws.append([estrato, pesos[pos]])
		ws.append([None, None])
		ws.append(self.styled_cells(["Número médio de docentes"], ws))
		ws.append(["ND", ND])

		ws = self.create_sheet("Autores")
		for row in dataframe_to_rows(self.autores, index=False, header=True):
			ws.append(row)

		lista_autores = []
		for pos, autor in enumerate(self.relatorios["Author"]):
			if autor.split(" ")[0] not in lista_autores:
				lista_autores.append(autor.split(" ")[0])
			else:
				encontrou = False
				for autor2 in self.relatorios["Author"]:
					if autor2.split(" ")[0] == autor.split(" ")[0] and encontrou == False:
						encontrou = True
						for pos, autor3 in enumerate(lista_autores):
							if autor3 == autor2.split(" ")[0]:
								lista_autores[pos] = f"{autor2.split(' ')[0]} {autor2.split(' ')[1]}"
				lista_autores.append(f"{autor.split(' ')[0]} {autor.split(' ')[1]}")
				
		indicadores_copy = []
		for tabela in self.indicadores:
			indicadores_copy.append(tabela.copy())

		for pos, tabela in enumerate(indicadores_copy):
			lista_nome = []
			for i in range(len(tabela.index)):
				lista_nome.append(lista_autores[pos])
			tabela["Nome Autor"] = lista_nome
			indicadores_copy[pos] = tabela

		indicadores_copy = pd.concat(indicadores_copy, ignore_index=True, sort=False)

		ws = self.create_sheet("Gráficos")
		graficos = Graficos(indicadores_copy, self.per_a1_a4, self.per_a1_a4_ae, lista_autores)
		ws = graficos.add_graphs(ws)

		ws = self.create_sheet("Resumo")
		resumo = pd.DataFrame(columns=["Nome", "Autores/artigo", "Irestrito", "Igeral", "Periódicos", "Anais", "A1-A4", "A1", "A2", "A3", "A4", "B1-B4", "B1", "B2", "B3", "B4", "Outros", "Periódicos A/E", "Anais A/E", "A1-A4 A/E", "A1 A/E", "A2 A/E", "A3 A/E", "A4 A/E", "B1-B4 A/E", "B1 A/E", "B2 A/E", "B3 A/E", "B4 A/E", "Outros A/E", "Periódicos A1-A4", "Periódicos A1-A4 com alunos/egressos"])
		posicoes = {"Irestrito": 15, "Igeral": 16, "Periódicos": 0, "Anais": 1, "A1-A4": 2, "A1": 3, "A2": 4, "A3": 5, "A4": 6, "B1-B4": 7, "B1": 8, "B2": 9, "B3": 10, "B4": 11, "Outros": 12}
		for pos, tabela in enumerate(self.indicadores):
			row = []
			row.append(lista_autores[pos])
			try:
				row.append(float(str(self.medias[pos].replace("Média de autores/artigo = ", ""))))
			except:
				row.append("")
			for key in posicoes.keys():
				row.append(tabela["Quantidade"][posicoes[key]])
			for key in posicoes.keys():
				if key != "Irestrito" and key != "Igeral":
					try:
						row.append(int(tabela["Quantidade com alunos/egressos"][posicoes[key]]))
					except:
						row.append(flaot(tabela["Quantidade com alunos/egressos"][posicoes[key]]))

			row.append(self.per_a1_a4[pos])
			row.append(self.per_a1_a4_ae[pos])
			resumo.loc[len(resumo)] = row

		row1 = []
		row2 = []
		
		for column in resumo.columns:
			ppg_total = 0
			if column != "Autores/artigo" and column != "Nome":
				for dado in resumo[column]:
					ppg_total += dado
				ppg_doc = ppg_total/ND

				ppg_total = round(ppg_total, 1)
				ppg_doc = round(ppg_doc, 1)
			elif column == "Nome":
				ppg_total = "PPGtotal"
				ppg_doc = "PPGdoc"
			else:
				ppg_total = "-"
				ppg_doc = "-"
			row1.append(ppg_total)
			row2.append(ppg_doc)
		
		resumo.loc[len(resumo)] = row1
		resumo.loc[len(resumo)] = row2
		
		ws.merge_cells('B1:D1')
		ws["A1"] = " "
		ws["B1"] = "Índices"
		ws.merge_cells('E1:Q1')
		ws["E1"] = "Publicações totais"
		ws.merge_cells('R1:AD1')
		ws["R1"] = "Publicações com alunos/egressos"
		ws.merge_cells('AE1:AF1')
		ws["AE1"] = " "
		
		resumo = pd.DataFrame(resumo)
		for row in dataframe_to_rows(resumo, index=False, header=True):
			ws.append(row)

		ws = self.create_sheet("Art|Prof")
		for row in dataframe_to_rows(self.todos, index=False, header=True):
			ws.append(row)


		result_df = self.artppg.apply(lambda x: x.astype(str).str.lower()).drop_duplicates(subset="Título")
		# artppg = artppg.drop_duplicates(subset="Título")
		self.artppg = self.artppg.loc[result_df.index]

		# citacoes = []
		# scopusModified = ScopusModified('2f8a856ea2c32c265b4c5a9895e6900d')
		# for pos, issn in enumerate(artppg["ISSN/SIGLA"]):
		# 	if artppg["Tipo"][pos] == "Periódico":
		# 		search = scopusModified.search(f"ISSN ({issn})")
		# 		scopus_id = search['scopus_id'][0]
		# 		cit = scopusModified.retrieve_citation(scopus_id_array=[scopus_id], year_range=[1990, 2020])
		# 		print(cit)
		# 		citacoes.append(cit)
		# 	else:
		# 		citacoes.append("-")

		# artppg["Citações"] = citacoes
		

		ws = self.create_sheet("Art|PPG")
		for row in dataframe_to_rows(self.artppg, index=False, header=True):
			ws.append(row)
		ws.append([""])
		for row in dataframe_to_rows(self.indicadores_geral, index=False, header=True):
			ws.append(self.styled_cells(row, ws))
		ws.append([None, None, None])
		ws.append([None, None, None])
		ws.append([None, None, None])

		ws = self.create_sheet("Anais|PPG")
		df = pd.DataFrame()
		df["Nome de Publicação"] = self.anais["Nome de Publicação"]
		df["Sigla"] = self.anais["SIGLA"]
		df["Qualis CC 2016"] = self.anais["Qualis CC 2016"]
		df["Qualis 2019"] = self.anais["Qualis 2019"]
		df["Quantidade"] = self.anais["Quantidade"]
		soma = 0
		for i in df["Quantidade"]:
			soma += i
		porcentagens = []
		for i in df["Quantidade"]:
			porcentagens.append(f"{round(100/soma * i, 1)}%")
		df["Porcentagem"] = porcentagens

		for row in dataframe_to_rows(df, index=False, header=True):
			ws.append(row)
		ws.append([None])
		for row in dataframe_to_rows(self.anais_metricas, index=False, header=True):
			ws.append(self.styled_cells(row, ws, pinta=False))
			
		graficos = Graficos_Anais_Periodicos(df.copy(), "Anais de Eventos Utilizados para Publicação")
		ws = graficos.add_graphs(ws)

		ws = self.create_sheet("Periódicos|PPG")
		df = pd.DataFrame()
		df["Nome de Publicação"] = self.periodicos["Nome de Publicação"]
		df["ISSN"] = self.periodicos["ISSN"]
		df["Qualis CC 2016"] = self.periodicos["Qualis CC 2016"]
		df["Qualis 2019"] = self.periodicos["Qualis 2019"]
		df["Scopus 2019"] = self.periodicos["Scopus 2019"]
		df["Quantidade"] = self.periodicos["Quantidade"]
		soma = 0
		for i in df["Quantidade"]:
			soma += i
		porcentagens = []
		for i in df["Quantidade"]:
			porcentagens.append(f"{round(100/soma * i, 1)}%")
		df["Porcentagem"] = porcentagens

		for row in dataframe_to_rows(df, index=False, header=True):
			ws.append(row)
		ws.append([None])
		for row in dataframe_to_rows(self.periodicos_metricas, index=False, header=True):
			ws.append(self.styled_cells(row, ws, pinta=False))

		graficos = Graficos_Anais_Periodicos(df.copy(), "Periódicos Utilizados para Publicação")
		ws = graficos.add_graphs(ws)

		for pos, autor in enumerate(self.relatorios["Author"]):
			row_count = 1
			ws = self.create_sheet(autor.split(" ")[0])
			for row in dataframe_to_rows(self.relatorios["Report"][pos], index=False, header=True): # Adiciona o dataframe ao sheet
				ws.append(row)
				row_count += 1

			ws.append([None, None, None])
			for row in dataframe_to_rows(self.indicadores[pos], index=False, header=True):
				ws.append(self.styled_cells(row, ws))
			ws.append([None, None, None])
			media = Cell(ws, column="A", row=1, value=self.medias[pos])
			media.font = Font(bold=True)
			ws.append([None, media])
			#ws.append([self.medias[pos]])
			ws.append([None, None, None])
			ws.append([None, None, None])
			ws.append([None, None, None])


		ws = self.create_sheet("Exceções")
		pos_artigos = False
		for pos, row in enumerate(dataframe_to_rows(self.excecoes, index=False, header=True)):
			if "Nome Periódico" in row and "Qualis 2019/ISSN Impresso" in row and "Qualis 2019/ISSN Online" in row:
				ws.append(self.styled_cells(row, ws, pinta=False))
				pos_artigos = True
			elif pos_artigos == True:
				ws.append(self.styled_cells(row, ws, pinta=False))
			else:
				ws.append(row)



	def altera_autores(self):
		lista_autores = []
		lista_egressos = []
		lista_alunos = []

		for autor in self.autores["Autor"]:
			lista_autores.append(autor)
		for egresso in self.lista_egressos:
			lista_egressos.append(egresso.name)
		for aluno in self.lista_alunos:
			lista_alunos.append(aluno.name)

		for ws in self.worksheets:
			for col in ws.columns:
				if (ws.title == "Art|Prof" and col[0].column_letter not in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']) or (ws.title != "Art|Prof" and col[0].column_letter not in ['A', 'B', 'C', 'D', 'E', 'F', 'G']):
					for pos, cell in enumerate(col):
						if cell.row != 1 and str(cell.value) != "" and str(cell.value) != " " and cell.value != None and str(cell.value) != 'nan':
							temp, cell.value = em_lista_autores(lista_autores, str(cell.value))
							temp, cell.value = em_lista_autores(lista_egressos, str(cell.value))
							temp, cell.value = em_lista_autores(lista_alunos, str(cell.value))


	def aplica_estilo(self):
		for ws in self.worksheets:
			for col in ws.columns:
				for cell in col:
					if cell.value != None and str(cell.value) != "nan" and "Média de autores" not in str(cell.value) and cell.value != "Número médio de docentes":
						bd = Side(border_style="thin", color="000000")
						cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)

			# Estilo dos nomes das colunas
			for cell in ws[1]:
				cell.font = Font(bold=True) # Negrito
				cell.alignment = Alignment(horizontal='center', vertical='center') # Centralizado

			if ws.title != 'Autores' and ws.title != "Exceções" and ws.title != "Resumo" and ws.title != "Anais|PPG" and ws.title != "Periódicos|PPG":
				# Centraliza todas as colunas menos a de título e de publicação
				for col in ws.columns:
					if ws.title != "Art|Prof":
						if col[0].column_letter not in 'CD':
							for cell in col:
								if cell.value != "Número médio de docentes":
									cell.alignment = Alignment(horizontal='center', vertical='center')
					else:
						if col[0].column_letter not in 'DE':
							for cell in col:
								cell.alignment = Alignment(horizontal='center', vertical='center')

			elif ws.title == "Anais|PPG" or ws.title =="Periódicos|PPG":
				for col in ws.columns:
					if col[0].column_letter != "A":
						for cell in col:
							cell.alignment = Alignment(horizontal='center', vertical='center')

			elif ws.title != "Exceções":
				for col in ws.columns:
					for cell in col:
						cell.alignment = Alignment(horizontal='center', vertical='center')

			if ws.title == "Resumo":
				for col in ws.columns:
					for pos, cell in enumerate(col):
						if pos >= ws.max_row -2:
							cell.font = Font(bold=True)
						if pos > 0:
							if pos == 1:
								cell.font = Font(bold=True)
							if cell.column in [3, 5, 7, 8, 9, 10, 11, 18, 20, 21, 22, 23, 24, 31, 32]:
								cell.font = Font(color='FFFAFA', bold=True)
						if pos == 1:
							if "A/E" in str(cell.value):
								cell.value = str(cell.value).replace(" A/E", "")



	def aplica_cores(self):
		for ws in self.worksheets:
			if ws.title != 'Autores' and ws.title != "Parâmetros" and ws.title != "Gráficos" and ws.title != "Resumo" and ws.title != "Anais|PPG" and ws.title != "Periódicos|PPG":
				if ws.title == "Art|Prof":
					tipo_column = "C"
					Q16_column = "G"
					Q20_column = "H"
					scopus_column = "I"
					# Q16_column = "GH"
					# Q20_column = "I"
				else:
					tipo_column = "B"
					Q16_column = "F"
					Q20_column = "G"
					scopus_column = "H"
					# Q16_column = "FG"
					# Q20_column = "H"

				
				for col in ws.iter_cols(min_col=None, max_col=None, min_row=None, max_row=None):
					for cell in col:
						if cell.value is None or str(cell.value) =="nan":
							cell.fill = PatternFill(fill_type='solid', start_color='FFFFFF', end_color='FFFFFF')
						
				for col in ws.columns:
					if (ws.title == "Art|Prof" and col[0].column_letter not in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']) or (ws.title != "Art|Prof" and col[0].column_letter not in ['A', 'B', 'C', 'D', 'E', 'F', 'G']):
						for row, cell in enumerate(col):
							for egresso in self.lista_egressos:
								if str(cell.value).lower() == egresso.name.lower():
									if ws.title == "Art|Prof":
										if egresso.period[str(int(ws.cell(row=row+1, column=2).value))[2:]] == True and int(str(int(ws.cell(row=row+1, column=2).value))[2:]) > egresso.egress_year:
											cell.fill = PatternFill(fill_type='solid', start_color='00FFFF', end_color='00FFFF')
											ws.cell(row=row+1, column=10).value = "X"
										elif egresso.period[str(int(ws.cell(row=row+1, column=2).value))[2:]] == True:
											cell.fill = PatternFill(fill_type='solid', start_color='F781F3', end_color='F781F3')
											ws.cell(row=row+1, column=10).value = "X"
									elif ws.title == "Art|PPG":
										if egresso.period[str(int(ws.cell(row=row+1, column=1).value))[2:]] == True and int(str(int(ws.cell(row=row+1, column=1).value))[2:]) > egresso.egress_year:
											cell.fill = PatternFill(fill_type='solid', start_color='00FFFF', end_color='00FFFF')
											ws.cell(row=row+1, column=10).value = "X"
										elif egresso.period[str(int(ws.cell(row=row+1, column=1).value))[2:]] == True:
											cell.fill = PatternFill(fill_type='solid', start_color='F781F3', end_color='F781F3')
											ws.cell(row=row+1, column=10).value = "X"
									else:
										if egresso.period[str(int(ws.cell(row=row+1, column=1).value))[2:]] == True and int(str(int(ws.cell(row=row+1, column=1).value))[2:]) > egresso.egress_year:
											cell.fill = PatternFill(fill_type='solid', start_color='00FFFF', end_color='00FFFF')
											ws.cell(row=row+1, column=9).value = "X"
										elif egresso.period[str(int(ws.cell(row=row+1, column=1).value))[2:]] == True:
											cell.fill = PatternFill(fill_type='solid', start_color='F781F3', end_color='F781F3')
											ws.cell(row=row+1, column=9).value = "X"
							for aluno in self.lista_alunos:
								if str(cell.value).lower() == aluno.name.lower():
									if ws.title == "Art|Prof":
										if aluno.period[str(ws.cell(row=row+1, column=2).value)[2:]] == True:
											cell.fill = PatternFill(fill_type='solid', start_color='F781F3', end_color='F781F3')
											ws.cell(row=row+1, column=10).value = "X"
									elif ws.title == "Art|PPG":
										if aluno.period[str(ws.cell(row=row+1, column=1).value)[2:]] == True:
											cell.fill = PatternFill(fill_type='solid', start_color='F781F3', end_color='F781F3')
											ws.cell(row=row+1, column=10).value = "X"
									else:
										if aluno.period[str(ws.cell(row=row+1, column=1).value)[2:]] == True:
											cell.fill = PatternFill(fill_type='solid', start_color='F781F3', end_color='F781F3')
											ws.cell(row=row+1, column=9).value = "X"

					if col[0].column_letter in tipo_column:
						for cell in col:
							if cell.value == "Periódico":
								cell.font = Font(color='FFFAFA')
								cell.fill = PatternFill(fill_type='solid', start_color='008000', end_color='008000')
							elif cell.value == "Anais":
								cell.fill = PatternFill(fill_type='solid', start_color='FFFF99', end_color='FFFF99')

					elif col[0].column_letter in Q16_column:
						for cell in col:
							if "A1" in str(cell.value) or "A2" in str(cell.value) or "B1" in str(cell.value):
								cell.font = Font(color='FFFAFA')
								cell.fill = PatternFill(fill_type='solid', start_color='008000', end_color='008000')
							elif "B2" in str(cell.value) or "B3" in str(cell.value) or "B4" in str(cell.value) or "B5" in str(cell.value):
								cell.fill = PatternFill(fill_type='solid', start_color='FFFF99', end_color='FFFF99')

					elif col[0].column_letter in Q20_column:
						for cell in col:
							if "A1" in str(cell.value) or "A2" in str(cell.value) or "A3" in str(cell.value) or "A4" in str(cell.value):
								cell.font = Font(color='FFFAFA')
								cell.fill = PatternFill(fill_type='solid', start_color='008000', end_color='008000')
							elif "B1" in str(cell.value) or "B2" in str(cell.value) or "B3" in str(cell.value) or "B4" in str(cell.value):
								cell.fill = PatternFill(fill_type='solid', start_color='FFFF99', end_color='FFFF99')

					elif col[0].column_letter in scopus_column:
						for cell in col:
							if "Q1" in str(cell.value) or "Q2" in str(cell.value):
								cell.font = Font(color='FFFAFA')
								cell.fill = PatternFill(fill_type='solid', start_color='008000', end_color='008000')
							elif "Q3" in str(cell.value) or "Q4" in str(cell.value):
								cell.fill = PatternFill(fill_type='solid', start_color='FFFF99', end_color='FFFF99')
			
			elif ws.title == "Gráficos":
				for i in range(7):
					for j in range(50):
						cell = ws.cell(row=j+1, column=i+1, value=None)
						cell.fill = PatternFill(fill_type='solid', start_color='FFFFFF', end_color='FFFFFF')

			elif ws.title == "Anais|PPG" or ws.title == "Periódicos|PPG":
				for col in ws.columns: 
					if col[0].column_letter in "CDE":
						for cell in col:
							if str(cell.value) in ["A1", "A2", "A3", "A4"]:
								cell.font = Font(color='FFFAFA')
								cell.fill = PatternFill(fill_type='solid', start_color='008000', end_color='008000')
							elif str(cell.value) in ["B1", "B2", "B3", "B4"]:
								cell.fill = PatternFill(fill_type='solid', start_color='FFFF99', end_color='FFFF99')
							elif "Q1" in str(cell.value) or "Q2" in str(cell.value):
								cell.font = Font(color='FFFAFA')
								cell.fill = PatternFill(fill_type='solid', start_color='008000', end_color='008000')
							elif "Q3" in str(cell.value) or "Q4" in str(cell.value):
								cell.fill = PatternFill(fill_type='solid', start_color='FFFF99', end_color='FFFF99')

			elif ws.title == "Resumo":
				verde = ["C", "E", "G", "H", "I", "J", "K", "R", "T", "U", "V", "W", "X", "AE", "AF"]
				branco = ["A", "B", "Q", "AD"]
				for col in ws.columns:
					for pos, cell in enumerate(col):
						if pos > 0:
							if col[1].column_letter in verde:
								cell.fill = PatternFill(fill_type='solid', start_color='00B050', end_color='008000')
							elif col[1].column_letter not in branco:
								cell.fill = PatternFill(fill_type='solid', start_color='FFFFCC', end_color='FFFF99')
			else:	
				for col in ws.columns:
					for cell in col:
						cell.alignment = Alignment(horizontal='center', vertical='center')
	def aplica_filtros(self):
		for ws in self.worksheets:
			if ws.title != "Parâmetros" and ws.title != "Art|Prof" and ws.title != "Gráficos" and ws.title != "Resumo" and ws.title != "Anais|PPG" and ws.title != "Periódicos|PPG":
				linhas = ws.dimensions[4:]
				linhas = int(linhas) -25
				dimensao = ws.dimensions[:4] + str(linhas)
				ws.auto_filter.ref = dimensao
			elif ws.title == "Anais|PPG" or ws.title == "Periódicos|PPG":
				linhas = ws.dimensions[4:]
				linhas = int(linhas) -6
				dimensao = ws.dimensions[:4] + str(linhas)
				ws.auto_filter.ref = dimensao

			elif ws.title != "Gráficos" and ws.title != "Resumo":
				ws.auto_filter.ref = ws.dimensions


	def aplica_dimensoes(self):
		for ws in self.worksheets:
			if ws.title == "Exceções":
				ws.column_dimensions['A'].width = 50
				ws.column_dimensions['B'].width = 50
				tamanho = len(self.excecoes["Nome Evento Canônico"][0])
				for i in self.excecoes["Nome Evento Canônico"]:
					if len(str(i)) > tamanho:
						tamanho = len(i)
				ws.column_dimensions['C'].width = tamanho

			elif ws.title == "Parâmetros":
				ws.column_dimensions['A'].width = 28
				ws.column_dimensions['B'].width = 15

			elif ws.title == "Gráficos":
				ws.column_dimensions['A'].width = 30
				ws.column_dimensions['B'].width = 30
				ws.column_dimensions['C'].width = 30
				ws.column_dimensions['D'].width = 30
				ws.column_dimensions['E'].width = 30
				ws.column_dimensions['F'].width = 30
				ws.column_dimensions['G'].width = 30
				for i in range(50):
					ws.row_dimensions[i+1].height = 50

			elif ws.title == "Anais|PPG":
				ws.column_dimensions['A'].width = 60
				ws.column_dimensions['B'].width = 15
				ws.column_dimensions['C'].width = 18
				ws.column_dimensions['D'].width = 18
				ws.column_dimensions['E'].width = 18
				ws.column_dimensions['F'].width = 15
				ws.column_dimensions['G'].width = 15

			elif ws.title == "Periódicos|PPG":
				ws.column_dimensions['A'].width = 60
				ws.column_dimensions['B'].width = 15
				ws.column_dimensions['C'].width = 18
				ws.column_dimensions['D'].width = 18
				ws.column_dimensions['E'].width = 18
				ws.column_dimensions['F'].width = 18
				ws.column_dimensions['G'].width = 15

			elif ws.title == "Resumo":
				ws.column_dimensions['A'].width = 15
				ws.column_dimensions['B'].width = 15
				columns = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", 
				"N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", 
				"AA", "AB", "AC", "AD"]
				for column in columns:
					ws.column_dimensions[column].width = 10
				ws.column_dimensions["AE"].width = 35
				ws.column_dimensions["AF"].width = 35

			elif ws.title != 'Autores':
				dic = {1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J', 11: 'K', 12: 'L', 13 : 'M', 14: 'N', 15: 'O', 16: 'P', 17: 'Q', 18: 'R', 19: 'S', 20: 'T', 21: 'U', 22: 'V', 23: 'W', 24: 'X', 25: 'Y'}
				# Define tamanho das colunas
				if ws.title != "Art|Prof" and ws.title != "Art|PPG":
					ws.column_dimensions['A'].width = 15
					ws.column_dimensions['B'].width = 15
					ws.column_dimensions['C'].width = 35
					ws.column_dimensions['D'].width = 35
					ws.column_dimensions['E'].width = 15
					ws.column_dimensions['F'].width = 18
					ws.column_dimensions['G'].width = 18
					ws.column_dimensions['H'].width = 18
					ws.column_dimensions['I'].width = 10
					ws.column_dimensions['J'].width = 30

					i = 11
					while i <= ws.max_column:
						ws.column_dimensions[dic[i]].width = 30
						i += 1
			
					ws.row_dimensions[1].height = 25 # Altura da linha dos nomes das colunas
				elif ws.title == "Art|PPG":
					ws.column_dimensions['A'].width = 15
					ws.column_dimensions['B'].width = 15
					ws.column_dimensions['C'].width = 35
					ws.column_dimensions['D'].width = 35
					ws.column_dimensions['E'].width = 15
					ws.column_dimensions['F'].width = 18
					ws.column_dimensions['G'].width = 18
					ws.column_dimensions['H'].width = 18
					ws.column_dimensions['I'].width = 18
					ws.column_dimensions['J'].width = 10
					ws.column_dimensions['K'].width = 30

					i = 12
					while i <= ws.max_column:
						ws.column_dimensions[dic[i]].width = 30
						i += 1
				else:
					ws.column_dimensions['A'].width = 35
					ws.column_dimensions['B'].width = 15
					ws.column_dimensions['C'].width = 15
					ws.column_dimensions['D'].width = 35
					ws.column_dimensions['E'].width = 35
					ws.column_dimensions['F'].width = 15
					ws.column_dimensions['G'].width = 18
					ws.column_dimensions['H'].width = 18
					ws.column_dimensions['I'].width = 18
					ws.column_dimensions['J'].width = 10
					ws.column_dimensions['K'].width = 30

					i = 12
					while i <= ws.max_column:
						ws.column_dimensions[dic[i]].width = 30
						i += 1
			
					ws.row_dimensions[1].height = 25 # Altura da linha dos nomes das colunas
			else:
				ws.column_dimensions['A'].width = 50
				ws.row_dimensions[1].height = 25 # Altura da linha dos nomes das colunas

	def salva_arquivo(self):
		self.save('lattes_qualis.xlsx') # Salva o arquivo
		for file in os.listdir("temp"):
			os.remove(f"temp/{file}") # Remove the file
		os.rmdir("temp") # Remove the temporary folder