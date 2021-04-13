import xml.etree.ElementTree as ET
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import jellyfish
import unidecode
from verifica_autores import trata_virgulas, em_lista_autores

class Autor():
	def __init__(self, nome, periodo, qualis_cc2016_file, qualis_e42016_file, qualis_xx2020_file, qualis_cc2016_eventos, qualis_xx2020_eventos, egressos, lista_autores):
		super(Autor, self).__init__()
		self.nome = nome
		self.periodo = periodo
		self.qualis_cc2016_file = qualis_cc2016_file
		self.qualis_e42016_file = qualis_e42016_file
		self.qualis_xx2020_file = qualis_xx2020_file

		self.qualis_cc2016_eventos = qualis_cc2016_eventos
		self.qualis_xx2020_eventos = qualis_xx2020_eventos

		self.egressos = egressos
		self.lista_autores = lista_autores
		
		self.dic_tipo = {"ARTIGO-PUBLICADO": "Periódico", "TRABALHO-EM-EVENTOS": "Anais"}
		self.info = {"Ano":[], "Tipo":[], "Título":[], "Nome de Publicação":[], "ISSN":[], "Qualis Eng. IV 2016":[], "Qualis CC 2016":[], "Qualis 2020":[]}
		self.sufixos = ['Jr.', 'Jr', 'Filho', 'Neto']
		# self.lista_autores = []
		self.qtd_autores = 0
		
		self.myroot = self.get_XML_file()
		self.fill_info()

	def get_XML_file(self):
		mytree = ET.parse(f"Curriculos Alunos/{self.nome}.xml")
		return mytree.getroot()

	def add_authors(self, pub):
		autores = []
		for i in pub:
			if 'NOME-COMPLETO-DO-AUTOR' in i.attrib.keys():
				self.qtd_autores += 1
				autor = i.attrib['NOME-COMPLETO-DO-AUTOR'] 
				autor = autor.title()
				
				autor = trata_virgulas(autor)

				self.lista_autores, autor = em_lista_autores(self.lista_autores, autor)

				autores.append(autor)
		
		for pos2, autor in enumerate(autores):
			if f"{pos2+1}º Autor" not in self.info:
				self.info[f"{pos2+1}º Autor"] = []
				for i in range(len(self.info["Título"])-1):
					self.info[f"{pos2+1}º Autor"].append("")

		autor_keys = []
		for key in self.info.keys():
			if "º Autor" in key:
				autor_keys.append(key)
		for pos3, key in enumerate(autor_keys):
			if pos3 <= (len(autores)-1):
				self.info[key].append(autores[pos3])
			else:
				self.info[key].append("")


	def add_data_artigo(self, pub, ano):
		self.info["Ano"].append(ano)
					
		self.info["Título"].append(pub[0].attrib['TITULO-DO-ARTIGO'])
		self.info["Nome de Publicação"].append(pub[1].attrib['TITULO-DO-PERIODICO-OU-REVISTA'])
		
		issn = pub[1].attrib['ISSN']
		self.info["ISSN"].append(issn[:4] + "-" +issn[4:])

		qualis_cc2016 = self.qualis_cc2016_file.loc[self.qualis_cc2016_file['ISSN'] == issn]
		try:
			qualis_cc2016 = qualis_cc2016.reset_index()
			qualis_cc2016 = qualis_cc2016["Estrato"][0]
		except:
			qualis_cc2016 = "-"

		self.info["Qualis CC 2016"].append(qualis_cc2016)

		qualis_e42016 = self.qualis_e42016_file.loc[self.qualis_e42016_file['ISSN'] == issn]
		try:
			qualis_e42016 = qualis_e42016.reset_index()
			qualis_e42016 = qualis_e42016["Estrato"][0]
		except:
			qualis_e42016 = "-"

		self.info["Qualis Eng. IV 2016"].append(qualis_e42016)

		qualis_cc2020 = self.qualis_xx2020_file.loc[self.qualis_xx2020_file['ISSN'] == issn]
		try:
			qualis_cc2020 = qualis_cc2020.reset_index()
			qualis_cc2020 = qualis_cc2020["Estrato"][0]
		except:
			qualis_cc2020 = "-"

		self.info["Qualis 2020"].append(qualis_cc2020)

	def add_data_evento(self, pub, ano):
		self.info["Ano"].append(ano)
		self.info["Título"].append(pub[0].attrib['TITULO-DO-TRABALHO'])
		self.info["ISSN"].append(" ")
		nome_evento = pub[1].attrib['NOME-DO-EVENTO']
		titulo_anais = pub[1].attrib['TITULO-DOS-ANAIS-OU-PROCEEDINGS']

		if nome_evento > titulo_anais:
			nome = nome_evento
		else:
			nome = titulo_anais

		qualis_cc2016 = "-"
		qualis_2020 = "-"

		for pos, nome_padrao in enumerate(self.qualis_cc2016_eventos['Nome Padrão']):
			if nome_padrao.lower() in nome_evento.lower():
				nome = nome_padrao
				qualis_cc2016 = self.qualis_cc2016_eventos['Qualis 2016'][pos]
				break
			elif nome_padrao.lower() in titulo_anais.lower():
				nome = nome_padrao
				qualis_cc2016 = self.qualis_cc2016_eventos['Qualis 2016'][pos]
				break
		self.info["Qualis CC 2016"].append(qualis_cc2016)

		for pos, nome_padrao in enumerate(self.qualis_xx2020_eventos['Nome Padrão']):
			if nome_padrao.lower() in nome_evento.lower():
				nome = nome_padrao
				qualis_2020 = self.qualis_xx2020_eventos['Qualis 2020'][pos]
				break
			elif nome_padrao.lower() in titulo_anais.lower():
				nome = nome_padrao
				qualis_2020 = self.qualis_xx2020_eventos['Qualis 2020'][pos]
				break

		self.info["Qualis 2020"].append(qualis_2020)

		self.info["Qualis Eng. IV 2016"].append("-")

		nome = nome.title()
		
		self.info["Nome de Publicação"].append(nome)


	def fill_info(self):
		publications_list = []
		for i in self.myroot[1]:
			if i.tag == "ARTIGOS-PUBLICADOS" or i.tag == "TRABALHOS-EM-EVENTOS":
				publications_list.append(i)

		for publications in publications_list:
			for pub in publications:
				if pub.tag == "TRABALHO-EM-EVENTOS":
					ano = int(pub[0].attrib['ANO-DO-TRABALHO'])
				else:
					ano = int(pub[0].attrib['ANO-DO-ARTIGO'])
				if ano >= 2017:
					# print(self.nome, periodo)
					if periodo[str(ano)[2:]] == True:
						if pub.tag == "TRABALHO-EM-EVENTOS":
							if pub[0].attrib['NATUREZA'] == 'COMPLETO':
								self.info["Tipo"].append(self.dic_tipo[pub.tag])
								self.add_data_evento(pub, ano)
								self.add_authors(pub)
						else:
							self.info["Tipo"].append(self.dic_tipo[pub.tag])
							self.add_data_artigo(pub, ano)
							self.add_authors(pub)


class ExcelFile(Workbook):
	def __init__(self, relatorios, autores, todos, medias):
		super(ExcelFile, self).__init__()
		self.relatorios = relatorios
		self.autores = autores
		self.todos = todos
		self.medias = medias
		self.add_info()
		self.aplica_estilo()
		# self.converte_valores()
		self.aplica_dimensoes()
		self.aplica_cores()
		self.aplica_filtros()


	def add_info(self):
		ws = self.active # Primeiro sheet
		ws.title =  'Autores'
		for row in dataframe_to_rows(self.autores, index=False, header=True):
			ws.append(row)

		ws = self.create_sheet("Art|Egresso")
		for row in dataframe_to_rows(self.todos, index=False, header=True):
			ws.append(row)

		artppg = self.todos.drop(columns="Nome")
		artppg = artppg.drop_duplicates(subset="Título")
		ws = self.create_sheet("Art|PPG")
		for row in dataframe_to_rows(artppg, index=False, header=True):
			ws.append(row)

		for pos, autor in enumerate(self.relatorios["Autor"]):
			ws = self.create_sheet(autor.split(" ")[0])
			for row in dataframe_to_rows(self.relatorios["Relatorio"][pos], index=False, header=True): # Adiciona o dataframe ao sheet
			    ws.append(row)
			ws.append([])
			ws.append(["", self.medias[pos]])

	def aplica_estilo(self):
		for ws in self.worksheets:
			# Estilo dos nomes das colunas
			for cell in ws[1]:
				cell.font = Font(bold=True) # Negrito
				cell.alignment = Alignment(horizontal='center', vertical='center') # Centralizado

			if ws.title != 'Autores':
				# Centraliza todas as colunas menos a de título e de publicação
				for col in ws.columns:
					if ws.title != "Art|Egresso":
						if col[0].column_letter not in 'CD':
							for cell in col:
								cell.alignment = Alignment(horizontal='center', vertical='center')
					else:
						if col[0].column_letter not in 'DE':
							for cell in col:
								cell.alignment = Alignment(horizontal='center', vertical='center')
			else:
				for col in ws.columns:
					for cell in col:
						cell.alignment = Alignment(horizontal='center', vertical='center')

	def aplica_cores(self):
		for ws in self.worksheets:
			if ws.title != 'Autores':
				if ws.title == "Art|Egresso":
					Q16_column = "GH"
					Q20_column = "I"
				else:
					Q16_column = "FG"
					Q20_column = "H"

				for col in ws.columns:
					if col[0].column_letter in Q16_column:
						for cell in col:
							if cell.value in ["A1", "A2", "B1"]:
								cell.font = Font(color='FFFAFA')
								cell.fill = PatternFill(fill_type='solid', start_color='008000', end_color='008000')
							elif cell.value in ["B2", "B3", "B4", "B5"]:
								cell.fill = PatternFill(fill_type='solid', start_color='FFFF99', end_color='FFFF99')

					elif col[0].column_letter in Q20_column:
						for cell in col:
							if cell.value in ["A1", "A2", "A3", "A4"]:
								cell.font = Font(color='FFFAFA')
								cell.fill = PatternFill(fill_type='solid', start_color='008000', end_color='008000')
							elif cell.value in ["B1", "B2", "B3", "B4"]:
								cell.fill = PatternFill(fill_type='solid', start_color='FFFF99', end_color='FFFF99')
			else:
				for col in ws.columns:
					for cell in col:
						cell.alignment = Alignment(horizontal='center', vertical='center')
	def aplica_filtros(self):
		for ws in self.worksheets:
			ws.auto_filter.ref = ws.dimensions

	def aplica_dimensoes(self):
		for ws in self.worksheets:
			if ws.title != 'Autores':
				dic = {1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J', 11: 'K', 12: 'L', 13 : 'M', 14: 'N', 15: 'O', 16: 'P', 17: 'Q', 18: 'R', 19: 'S', 20: 'T', 21: 'U', 22: 'V', 23: 'W', 24: 'X', 25: 'Y'}
				# Define tamanho das colunas
				if ws.title != "Art|Egresso":
					ws.column_dimensions['A'].width = 15
					ws.column_dimensions['B'].width = 15
					ws.column_dimensions['C'].width = 35
					ws.column_dimensions['D'].width = 35
					ws.column_dimensions['E'].width = 15
					ws.column_dimensions['F'].width = 25
					ws.column_dimensions['G'].width = 25
					ws.column_dimensions['H'].width = 25
					ws.column_dimensions['I'].width = 30

					i = 10
					while i <= ws.max_column:
						ws.column_dimensions[dic[i]].width = 30
						i += 1
			
					ws.row_dimensions[1].height = 25 # Altura da linha dos nomes das colunas
				else:
					ws.column_dimensions['A'].width = 35
					ws.column_dimensions['B'].width = 15
					ws.column_dimensions['C'].width = 15
					ws.column_dimensions['D'].width = 35
					ws.column_dimensions['E'].width = 35
					ws.column_dimensions['F'].width = 15
					ws.column_dimensions['G'].width = 25
					ws.column_dimensions['H'].width = 25
					ws.column_dimensions['I'].width = 25
					ws.column_dimensions['J'].width = 30

					i = 11
					while i <= ws.max_column:
						ws.column_dimensions[dic[i]].width = 30
						i += 1
			
					ws.row_dimensions[1].height = 25 # Altura da linha dos nomes das colunas
			else:
				ws.column_dimensions['A'].width = 50
				ws.row_dimensions[1].height = 25 # Altura da linha dos nomes das colunas

	def salva_arquivo(self):
		self.save('lattes_qualis_alunos.xlsx') # Salva o arquivo



if __name__ == '__main__':
	alunos = pd.read_csv("Planilha - Levantamento alunos ativos.CSV", sep=";", encoding='iso-8859-1')

	qualis_cc2016_file = pd.read_csv("QualisCC_2013_2016.csv", sep=";", encoding='iso-8859-1')
	qualis_e42016_file = pd.read_csv("QualisEng.IV_2013_2016.csv", sep=";", encoding='iso-8859-1')
	qualis_xx2020_file = pd.read_csv("QualisXX_2020.csv", sep=";", encoding='iso-8859-1')

	qualis_cc2016_eventos = pd.read_csv("QualisCC_eventos_2016.csv", sep=";", encoding='iso-8859-1')
	for pos, i in enumerate(qualis_cc2016_eventos['Nome Padrão']):
		qualis_cc2016_eventos['Nome Padrão'][pos] = str(qualis_cc2016_eventos['Nome Padrão'][pos]).lower()

	qualis_xx2020_eventos = pd.read_csv("QualisXX_eventos_2020.csv", sep=";", encoding='iso-8859-1')
	for pos, i in enumerate(qualis_xx2020_eventos['Nome Padrão']):
		qualis_xx2020_eventos['Nome Padrão'][pos] = str(qualis_xx2020_eventos['Nome Padrão'][pos]).lower()

	for i in range(len(qualis_cc2016_file["ISSN"])):
		qualis_cc2016_file["ISSN"][i] = qualis_cc2016_file["ISSN"][i].replace("-", "")
	for i in range(len(qualis_e42016_file["ISSN"])):
		qualis_e42016_file["ISSN"][i] = qualis_e42016_file["ISSN"][i].replace("-", "")
	for i in range(len(qualis_xx2020_file["ISSN"])):
		qualis_xx2020_file["ISSN"][i] = qualis_xx2020_file["ISSN"][i].replace("-", "")


	relatorios = {'Autor':[], 'Relatorio':[]}
	lista_autores = []
	todos_lista = []
	medias_autores = []
	for pos, aluno in enumerate(alunos["Aluno"]):
			if str(aluno) != "nan" and str(aluno) != "Adson Marques da Silva Esteves" and str(aluno) != "Angelica Karize Viecelli":
				periodo = {"17": False, "18": False, "19": False, "20": False}
				if str(alunos["Ingresso"][pos]) != "nan":
					ano_ingresso = int(str(alunos["Ingresso"][pos])[7:])
					if str(ano_ingresso) in periodo:
						periodo[str(ano_ingresso)] = True
					ano_limite = int("20")
					for i in range(1, ano_limite+1):
						if str(ano_ingresso + i) in periodo:
							periodo[str(ano_ingresso + i)] = True

					ano_egresso = str(alunos["Defesa"][pos])
					if ano_egresso != '' and ano_egresso != "nan":
						ano_egresso = int(ano_egresso[7:])
						if str(ano_egresso) in periodo:
							periodo[str(ano_egresso)] = True
						for i in range(1, 6):
							if str(ano_egresso + i) in periodo:
								periodo[str(ano_egresso + i)] = True

					#====================================================

					autor = Autor(aluno, periodo, qualis_cc2016_file, qualis_e42016_file, qualis_xx2020_file, qualis_cc2016_eventos, qualis_xx2020_eventos, alunos, lista_autores)
					lista_autores = autor.lista_autores

					relatorios['Autor'].append(aluno)
					relatorios['Relatorio'].append(pd.DataFrame(autor.info))
					
					# for nome in autor.lista_autores:
					# 	if nome not in lista_autores:
					# 		lista_autores.append(nome)

					info_df = pd.DataFrame(autor.info)
					lista_nome = [autor.nome for i in range(len(info_df))]
					info_df.insert(loc=0, column='Nome', value=lista_nome)
					todos_lista.append(info_df)

					try:
						medias_autores.append(f"Média de autores/artigo = {autor.qtd_autores/len(pd.DataFrame(autor.info)):.2f}")
					except ZeroDivisionError:
						medias_autores.append(f"")
					
	dic_autores = {"Autor":lista_autores}
	todos_df = pd.concat(todos_lista, ignore_index=True, sort=False)
	excel = ExcelFile(relatorios, pd.DataFrame(dic_autores), todos_df, medias_autores)
	excel.salva_arquivo()

	
	