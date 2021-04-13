from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

import pandas as pd
import re
import os

from _Classes.PyscopusModified import ScopusModified

from _Funções_e_Valores.verify_authors import search_authors_list, treat_exceptions
from _Funções_e_Valores.values import ND, EXCEL_FILE_NAME, HAS_EVENTS, FILE

from _Classes.Graph import Graphs, Graphs_Proceedings_Journals


class ExcelFile(Workbook):
	def __init__(self, data):
		super(ExcelFile, self).__init__()
		self.reports = data.reports
		self.authors = pd.DataFrame(data.authors_dict)
		self.art_prof = data.art_prof
		self.artppg = data.artppg
		self.averages = data.authors_average

		self.irestritos_2016 = data.irestritos_2016
		self.igerais_2016 = data.igerais_2016
		self.irestritos_2019 = data.irestritos_2019
		self.igerais_2019 = data.igerais_2019

		self.authors_indicators_2016 = data.authors_indicators_2016
		self.authors_indicators_2019 = data.authors_indicators_2019
		self.general_indicators_2016 = data.general_indicators_2016
		self.general_indicators_2019 = data.general_indicators_2019

		self.authors_indicators_2016_journals = data.authors_indicators_2016_journals
		self.authors_indicators_2019_journals = data.authors_indicators_2019_journals
		self.general_indicators_2016_journals = data.general_indicators_2016_journals
		self.general_indicators_2019_journals = data.general_indicators_2019_journals

		self.authors_indicators_2016_proceedings = data.authors_indicators_2016_proceedings
		self.authors_indicators_2019_proceedings = data.authors_indicators_2019_proceedings
		self.general_indicators_2016_proceedings = data.general_indicators_2016_proceedings
		self.general_indicators_2019_proceedings = data.general_indicators_2019_proceedings

		self.egress_list = data.egress_list
		self.students_list = data.students_list
		self.exceptions = data.exceptions

		self.journals_upperstrata_2019 = data.journals_upperstrata_2019
		self.journals_upperstrata_SE_2019 = data.journals_upperstrata_SE_2019
		self.journals_upperstrata_2016 = data.journals_upperstrata_2016
		self.journals_upperstrata_SE_2016 = data.journals_upperstrata_SE_2016

		self.journals = data.journals
		self.proceedings = data.proceedings

		self.journal_metrics_2019 = data.journal_metrics_2019
		self.journal_metrics_2016 = data.journal_metrics_2016
		self.proceedings_metrics_2019 = data.proceedings_metrics_2019
		self.proceedings_metrics_2016 = data.proceedings_metrics_2016

		for pos, egress in enumerate(self.egress_list):
			self.egress_list[pos].name = treat_exceptions(egress.name.strip())
		for pos, student in enumerate(self.students_list):
			self.students_list[pos].name = treat_exceptions(student.name.strip())

		self.add_info()
		# self.altera_authors()
		self.apply_style()
		# self.converte_valores()
		self.apply_dimensions()
		self.apply_colors()
		self.apply_filters()


	def styled_cells(self, data, ws, paint=True, qualis_year=None): # Apply a specific style to the cell
		for c in data:
			c = Cell(ws, column="A", row=1, value=c)
			if c.value != None and str(c.value) != "nan":
				if c.value == "Porcentagem alunos/egressos":
					c.value = "Porcentagem" # Change the name of the cell
				if data[0] in ["Periódicos", "A1-A4", "A1", "A2", "A3", "A4", "Irestrito",  "Irestrito Periódicos",  "Irestrito Anais"]: # Check the cell column
					c.font = Font(color='FFFAFA') # Change the color of the text (to white)
					c.fill = PatternFill(fill_type='solid', start_color='00B050', end_color='00B050') # Change the background color of the cell (to green)
				elif qualis_year == "2016" and data[0] in ["A1-B1", "B1"]:
					c.font = Font(color='FFFAFA') # Change the color of the text (to white)
					c.fill = PatternFill(fill_type='solid', start_color='00B050', end_color='00B050') # Change the background color of the cell (to green)
				elif data[0] != "Outros" and data[0] != "Número médio de docentes" and paint == True: # Check the cell column
					c.fill = PatternFill(fill_type='solid', start_color='FFFFCC', end_color='FFFFCC') # Change the background color of the cell (simmilar to yellow)
				if c.value in ["Tipo/Qualis CC 2016", "Tipo/Qualis 2019", "Quantidade", "Porcentagem", "Quantidade com alunos/egressos", "% Alunos/Egressos", "Índice", "Acumulado", "Média por docente", "Número médio de docentes", "Nome Periódico", "Qualis 2019/ISSN Impresso", "Qualis 2019/ISSN Online", "Métrica CC 2016", "Métrica 2019", "Qtd.", "Qtd. %", "Periódicos e Anais - Qualis 2016", "Periódicos e Anais - Qualis 2019", "Periódicos - Qualis 2016", "Periódicos - Qualis 2019", "Anais - Qualis 2016", "Anais - Qualis 2019"]:
					c.font = Font(bold=True) # Set the text bold
				if c.value != "Número médio de docentes" and c.value != "Periódicos e Anais - Qualis 2016" and c.value != "Periódicos e Anais - Qualis 2019":
					bd = Side(border_style="thin", color="000000") # Black border
					c.border = Border(left=bd, top=bd, right=bd, bottom=bd) # Set the border
					c.alignment = Alignment(horizontal='center', vertical='center') # Center alignment
			yield c

	def add_parameters_sheet(self):
		ws = self.active # First sheet
		ws.title = 'Parâmetros'
		ws.append(["Estrato", "Peso"]) # Add two columns

		# Columns values:
		strata_list = ["A1", "A2", "A3", "A4", "B1", "B2", "B3", "B4", ]
		weights = [1.000, 0.875, 0.750, 0.625, 0.500, 0.200, 0.100, 0.050, ]
		
		for pos, strata in enumerate(strata_list):
			ws.append([strata, weights[pos]]) # Add the values to the columns

		ws.append([None, None]) # Blank line
		ws.append(self.styled_cells(["Número médio de docentes"], ws)) # Add ND
		ws.append(["ND", ND])

	def add_indicators_sheet(self):
		ws = self.create_sheet("Indicadores")
		ws.append(["Qualis 2013-2016"])
		ws.append(["Irestrito", "Acumulado", "Médio", None, "Igeral", "Acumulado", "Médio"]) # Add two columns

		for pos, item in enumerate(self.irestritos_2016.items()):
			item = list(item)

			mean = round(item[1]/ND, 2)
			mean = str(mean)
			mean = mean.replace('.', ',')
			item.append(mean)
			item.append(None)

			item.append(list(self.igerais_2016.items())[pos][0])
			item.append(list(self.igerais_2016.items())[pos][1])
			mean_2 = float(str(list(self.igerais_2016.items())[pos][1]).replace(',', '.'))
			mean_2 = round(mean_2/ND, 2)
			mean_2 = str(mean_2).replace('.', ',')
			item.append(mean_2)
			ws.append(item)

		ws.append([None, None, None, None, None, None, None])
		ws.append(["Qualis 2017-2020"])
		ws.append(["Irestrito", "Acumulado", "Médio", None, "Igeral", "Acumulado", "Médio"]) # Add two columns
		for pos, item in enumerate(self.irestritos_2019.items()):
			item = list(item)

			mean = round(item[1]/ND, 2)
			mean = str(mean)
			mean = mean.replace('.', ',')
			item.append(mean)
			item.append(None)

			item.append(list(self.igerais_2019.items())[pos][0])
			item.append(list(self.igerais_2019.items())[pos][1])
			mean_2 = float(str(list(self.igerais_2019.items())[pos][1]).replace(',', '.'))
			mean_2 = round(mean_2/ND, 2)
			mean_2 = str(mean_2).replace('.', ',')
			item.append(mean_2)
			ws.append(item)


	def add_authors_sheet(self):
		ws = self.create_sheet("Autores") # Sheet with a list of authors
		self.authors = self.authors.rename(columns={"Author": "Autor"})
		for row in dataframe_to_rows(self.authors, index=False, header=True): # Add the dataframe rows to the sheet
			ws.append(row)

	def get_authors_list(self): # Get a list with the authors names
		authors_list = []
		for pos, author in enumerate(self.reports["Author"]):
			if FILE == "EGRESSOS 2017-2020":
				first_name = author.split(" ")[0]
				second_name = author.split(" ")[1]
				if len(second_name) < 3:
					second_name = second_name + " " + author.split(" ")[2]
		
				authors_list.append(first_name + " " + second_name)

			else:
				if author.split(" ")[0] not in authors_list: # Add the author first name to the list if its not already in there
					authors_list.append(author.split(" ")[0])
				else:
					# If there's already an author with this name both of them will also carry their second name
					found = False
					for author2 in self.reports["Author"]:
						if author2.split(" ")[0] == author.split(" ")[0] and found == False:
							found = True
							for pos, autor3 in enumerate(authors_list):
								if autor3 == author2.split(" ")[0]:
									authors_list[pos] = f"{author2.split(' ')[0]} {author2.split(' ')[1]}"
					authors_list.append(f"{author.split(' ')[0]} {author.split(' ')[1]}")

		return authors_list

	def add_graphs_sheet(self, authors_list):
		authors_indicators_2016_copy = [] 
		for table in self.authors_indicators_2016:
			authors_indicators_2016_copy.append(table.copy()) # Creates a copy of each table and add to the indicators_copy list

		for pos, table in enumerate(authors_indicators_2016_copy):
			name_list = []
			for i in range(len(table.index)):
				name_list.append(authors_list[pos]) 
			table["Nome Autor"] = name_list # Add a list of authors names to the table 
			authors_indicators_2016_copy[pos] = table 

		authors_indicators_2016_copy = pd.concat(authors_indicators_2016_copy, ignore_index=True, sort=False)

		authors_indicators_2019_copy = [] 
		for table in self.authors_indicators_2019:
			authors_indicators_2019_copy.append(table.copy()) # Creates a copy of each table and add to the indicators_copy list

		for pos, table in enumerate(authors_indicators_2019_copy):
			name_list = []
			for i in range(len(table.index)):
				name_list.append(authors_list[pos]) 
			table["Nome Autor"] = name_list # Add a list of authors names to the table
			authors_indicators_2019_copy[pos] = table 

		authors_indicators_2019_copy = pd.concat(authors_indicators_2019_copy, ignore_index=True, sort=False)

		ws = self.create_sheet("Gráficos Q2016") # Creates the graphs sheet
		graphs_2016 = Graphs(authors_indicators_2016_copy, self.journals_upperstrata_2016, self.journals_upperstrata_SE_2016, authors_list, "CC 2016", "temp") # Generates the graphs
		ws = graphs_2016.add_graphs(ws) # Add the graphs

		ws = self.create_sheet("Gráficos Q2019") # Creates the graphs sheet
		graphs_2019 = Graphs(authors_indicators_2019_copy, self.journals_upperstrata_2019, self.journals_upperstrata_2019, authors_list, "2019", "temp2") # Generates the graphs
		ws = graphs_2019.add_graphs(ws) # Add the graphs


	def build_general_summary(self, ws, authors_list, qualis_year):
		ws.append(self.styled_cells([f"Periódicos e Anais - Qualis {qualis_year}", None, None, None, None, None, 
			None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, 
			None, None, None, None, None, None], ws, paint=False))
		ws.merge_cells('A1:D1')

		if qualis_year == "2016":
			indicators = self.authors_indicators_2016
			summary = pd.DataFrame(columns=["Nome", "Autores/Artigo", "Irestrito", "Igeral", "Periódicos", "Anais", "A1-B1", "A1", "A2", "B1", "B2-B5", "B2", "B3", "B4",  "B5", "Outros", "Periódicos A/E", "Anais A/E", "A1-B1 A/E", "A1 A/E", "A2 A/E", "B1 A/E", "B2-B5 A/E", "B2 A/E", "B3 A/E", "B4 A/E", "B5 A/E", "Outros A/E", "Periódicos A1-B1", "Periódicos A1-B1 com alunos/egressos"])
			positions = {"Irestrito": 14, "Igeral": 15, "Periódicos": 0, "Anais": 1, "A1-B1": 2, "A1": 3, "A2": 4, "B1":5, "B2-B5": 6, "B2": 7, "B3": 8, "B4": 9, "B5": 10, "Outros": 11}
		elif qualis_year == "2019":
			indicators = self.authors_indicators_2019
			summary = pd.DataFrame(columns=["Nome", "Autores/Artigo", "Irestrito", "Igeral", "Periódicos", "Anais", "A1-A4", "A1", "A2", "A3", "A4", "B1-B4", "B1", "B2", "B3", "B4", "Outros", "Periódicos A/E", "Anais A/E", "A1-A4 A/E", "A1 A/E", "A2 A/E", "A3 A/E", "A4 A/E", "B1-B4 A/E", "B1 A/E", "B2 A/E", "B3 A/E", "B4 A/E", "Outros A/E", "Periódicos A1-A4", "Periódicos A1-A4 com alunos/egressos"])
			positions = {"Irestrito": 15, "Igeral": 16, "Periódicos": 0, "Anais": 1, "A1-A4": 2, "A1": 3, "A2": 4, "A3": 5, "A4": 6, "B1-B4": 7, "B1": 8, "B2": 9, "B3": 10, "B4": 11, "Outros": 12}

		for pos, table in enumerate(indicators):
			row = []
			row.append(authors_list[pos])
			try:
				average = str(self.averages[pos]).replace("Média de autores/artigo = ", "")
				row.append(float(average))
			except:
				row.append("")
			for key in positions.keys():
				row.append(table["Quantidade"][positions[key]])
			for key in positions.keys():
				if key != "Irestrito" and key != "Igeral":
					try:
						row.append(int(table["Quantidade com alunos/egressos"][positions[key]]))
					except:
						row.append(flaot(table["Quantidade com alunos/egressos"][positions[key]]))

			if qualis_year == "2016":
				row.append(self.journals_upperstrata_2016[pos])
				row.append(self.journals_upperstrata_SE_2016[pos])
			elif qualis_year == "2019":
				row.append(self.journals_upperstrata_2019[pos])
				row.append(self.journals_upperstrata_SE_2019[pos])

			summary.loc[len(summary)] = row

		row1 = []
		row2 = []
		
		for column in summary.columns:
			total_ppg = 0
			if column != "Autores/Artigo" and column != "Nome" and column != "Irestrito" and column != "Igeral":
				for data in summary[column]:
					total_ppg += data
				doc_ppg = total_ppg/ND

				total_ppg = round(total_ppg, 1)
				doc_ppg = round(doc_ppg, 1)
			elif column == "Nome":
				total_ppg = "PPGtotal"
				doc_ppg = "PPGdoc"
			elif column == "Irestrito":
				if qualis_year == "2016":
					total_ppg = self.irestritos_2016['Total sem trava']
				elif qualis_year == "2019":
					total_ppg = self.irestritos_2019['Total sem trava']
				doc_ppg = total_ppg/ND

				total_ppg = round(total_ppg, 1)
				doc_ppg = round(doc_ppg, 1)
			elif column == "Igeral":
				if qualis_year == "2016":
					total_ppg = self.igerais_2016['Total sem trava']
				elif qualis_year == "2019":
					total_ppg = self.igerais_2019['Total sem trava']
				doc_ppg = total_ppg/ND

				total_ppg = round(total_ppg, 1)
				doc_ppg = round(doc_ppg, 1)
			else:
				total_ppg = "-"
				doc_ppg = "-"
			row1.append(total_ppg)
			row2.append(doc_ppg)
		
		summary.loc[len(summary)] = row1
		summary.loc[len(summary)] = row2
		
		ws.merge_cells('B2:D2')
		ws["A2"] = " "
		ws["B2"] = "Índices"
		if qualis_year == "2016":
			ws.merge_cells('E2:P2')
			ws["E2"] = "Publicações totais"
			ws.merge_cells('Q2:AB2')
			ws["Q2"] = "Publicações com alunos/egressos"
			ws.merge_cells('AC2:AD2')
			ws["AC2"] = " "
		else:
			ws.merge_cells('E2:Q2')
			ws["E2"] = "Publicações totais"
			ws.merge_cells('R2:AD2')
			ws["R2"] = "Publicações com alunos/egressos"
			ws.merge_cells('AE2:AF2')
			ws["AE2"] = " "


		bd = Side(border_style="thin", color="000000") # Black border
		row = list(ws.rows)[1]
		for pos, cell in enumerate(row):
			cell.border = Border(left=bd, top=bd, right=bd, bottom=bd) # Set the border
		
		summary = pd.DataFrame(summary)
		rows_count = 2 # The title + the first row
		for row in dataframe_to_rows(summary, index=False, header=True):
			ws.append(row)
			rows_count += 1

		return (ws, rows_count)

	def build_separated_summary(self, ws, authors_list, qualis_year, rows_count, indicators, pub_type):
		indicators = indicators
		if qualis_year == "2016":
			summary = pd.DataFrame(columns=["Nome", "Autores/Artigo", "Irestrito", "Igeral", "A1-B1", "A1", "A2", "B1", "B2-B5", "B2", "B3", "B4",  "B5", "Outros", "A1-B1 A/E", "A1 A/E", "A2 A/E", "B1 A/E", "B2-B5 A/E", "B2 A/E", "B3 A/E", "B4 A/E", "B5 A/E", "Outros A/E"])
			positions = {"Irestrito": 12, "Igeral": 13, "A1-B1": 0, "A1": 1, "A2": 2, "B1":3, "B2-B5": 4, "B2": 5, "B3": 6, "B4": 7, "B5": 8, "Outros": 9}
		elif qualis_year == "2019":
			summary = pd.DataFrame(columns=["Nome", "Autores/Artigo", "Irestrito", "Igeral", "A1-A4", "A1", "A2", "A3", "A4", "B1-B4", "B1", "B2", "B3", "B4", "Outros", "A1-A4 A/E", "A1 A/E", "A2 A/E", "A3 A/E", "A4 A/E", "B1-B4 A/E", "B1 A/E", "B2 A/E", "B3 A/E", "B4 A/E", "Outros A/E"])
			positions = {"Irestrito": 13, "Igeral": 14, "A1-A4": 0, "A1": 1, "A2": 2, "A3": 3, "A4": 4, "B1-B4": 5, "B1": 6, "B2": 7, "B3": 8, "B4": 9, "Outros": 10}

		for pos, table in enumerate(indicators):
			row = []
			row.append(authors_list[pos])
			try:
				average = str(self.averages[pos]).replace("Média de autores/artigo = ", "")
				row.append(float(average))
			except:
				row.append("")
			for key in positions.keys():
				row.append(table["Quantidade"][positions[key]])
			for key in positions.keys():
				if key != "Irestrito" and key != "Igeral":
					try:
						row.append(int(table["Quantidade com alunos/egressos"][positions[key]]))
					except:
						row.append(flaot(table["Quantidade com alunos/egressos"][positions[key]]))
			summary.loc[len(summary)] = row

		row1 = []
		row2 = []
		
		for column in summary.columns:
			total_ppg = 0
			if column != "Autores/Artigo" and column != "Nome" and column != "Irestrito" and column != "Igeral":
				for data in summary[column]:
					total_ppg += data
				doc_ppg = total_ppg/ND

				total_ppg = round(total_ppg, 1)
				doc_ppg = round(doc_ppg, 1)
			elif column == "Nome":
				total_ppg = "PPGtotal"
				doc_ppg = "PPGdoc"
			elif column == "Irestrito":
				if qualis_year == "2016":
					if pub_type == "journals":
						total_ppg = self.irestritos_2016['Periódicos']
					elif pub_type == "proceedings":
						total_ppg = self.irestritos_2016['Anais sem trava']
				elif qualis_year == "2019":
					if pub_type == "journals":
						total_ppg = self.irestritos_2019['Periódicos']
					elif pub_type == "proceedings":
						total_ppg = self.irestritos_2019['Anais sem trava']

				doc_ppg = total_ppg/ND
				total_ppg = round(total_ppg, 1)
				doc_ppg = round(doc_ppg, 1)
			elif column == "Igeral":
				if qualis_year == "2016":
					if pub_type == "journals":
						total_ppg = self.igerais_2016['Periódicos']
					elif pub_type == "proceedings":
						total_ppg = self.igerais_2016['Anais sem trava']
				elif qualis_year == "2019":
					if pub_type == "journals":
						total_ppg = self.igerais_2019['Periódicos']
					elif pub_type == "proceedings":
						total_ppg = self.igerais_2019['Anais sem trava']

				doc_ppg = total_ppg/ND
				total_ppg = round(total_ppg, 1)
				doc_ppg = round(doc_ppg, 1)
			else:
				total_ppg = "-"
				doc_ppg = "-"
			row1.append(total_ppg)
			row2.append(doc_ppg)
		
		summary.loc[len(summary)] = row1
		summary.loc[len(summary)] = row2
		
		ws.merge_cells(f'B{rows_count+1}:D{rows_count+1}')
		ws[f"A{rows_count+1}"] = " "
		ws[f"B{rows_count+1}"] = "Índices"
		if qualis_year == "2016":
			ws.merge_cells(f'E{rows_count+1}:N{rows_count+1}')
			ws[f"E{rows_count+1}"] = "Publicações totais"
			ws.merge_cells(f'O{rows_count+1}:X{rows_count+1}')
			ws[f"O{rows_count+1}"] = "Publicações com alunos/egressos"
			
			bd = Side(border_style="thin", color="000000") # Black border
			row = list(ws.rows)[rows_count]
			for pos, cell in enumerate(row):
				if pos < 24:
					cell.border = Border(left=bd, top=bd, right=bd, bottom=bd) # Set the border
		else:
			ws.merge_cells(f'E{rows_count+1}:O{rows_count+1}')
			ws[f"E{rows_count+1}"] = "Publicações totais"
			ws.merge_cells(f'P{rows_count+1}:Z{rows_count+1}')
			ws[f"P{rows_count+1}"] = "Publicações com alunos/egressos"

			bd = Side(border_style="thin", color="000000") # Black border
			row = list(ws.rows)[rows_count]
			for pos, cell in enumerate(row):
				if pos < 26:
					cell.border = Border(left=bd, top=bd, right=bd, bottom=bd) # Set the border
		
		summary = pd.DataFrame(summary)
		rows_count += 1 # The first row
		for row in dataframe_to_rows(summary, index=False, header=True):
			ws.append(row)
			rows_count += 1

		return (ws, rows_count)

	def add_summary_sheet(self, authors_list, qualis_year):
		ws = self.create_sheet(f"Resumo Q{qualis_year}") # Creates the summary sheet
		ws, rows_count = self.build_general_summary(ws, authors_list, qualis_year)	
		self.summary_size = rows_count

		ws.append([None, None, None, None, None, None, None, None, None, None, None, 
			None, None, None, None, None, None, None, None, None, None, None, None, 
			None, None, None, None, None, None, None])
		ws.append([None, None, None, None, None, None, None, None, None, None, None, 
			None, None, None, None, None, None, None, None, None, None, None, None, 
			None, None, None, None, None, None, None])
		ws.append(self.styled_cells([f"Periódicos - Qualis {qualis_year}", None, None, None, None, None, None, None, 
			None, None, None, None, None, None, None, None, None, None, None, None, 
			None, None, None, None, None, None, None], ws, paint=False))
		
		ws.merge_cells(f'A{rows_count+3}:D{rows_count+3}')
		bd = Side(border_style="thin", color="ffffff")
		ws[f"A{rows_count+2}"].border = Border(left=bd, top=bd, right=bd, bottom=bd)
		ws[f"A{rows_count+3}"].border = Border(left=bd, top=bd, right=bd, bottom=bd)
		ws[f"B{rows_count+3}"].border = Border(left=bd, top=bd, right=bd, bottom=bd)
		ws[f"C{rows_count+3}"].border = Border(left=bd, top=bd, right=bd, bottom=bd)
		ws[f"D{rows_count+3}"].border = Border(left=bd, top=bd, right=bd, bottom=bd)

		rows_count += 3

		if qualis_year == "2016":
			indicators = self.authors_indicators_2016_journals
		else:
			indicators = self.authors_indicators_2019_journals
		ws, rows_count = self.build_separated_summary(ws, authors_list, qualis_year, rows_count, indicators, pub_type="journals")

		ws.append([None, None, None, None, None, None, None, None, None, None, None, 
			None, None, None, None, None, None, None, None, None, None, None, None, 
			None, None, None, None, None, None, None])
		ws.append([None, None, None, None, None, None, None, None, None, None, None, 
			None, None, None, None, None, None, None, None, None, None, None, None, 
			None, None, None, None, None, None, None])
		ws.append(self.styled_cells([f"Anais - Qualis {qualis_year}", None, None, None, None, None, None, None, 
			None, None, None, None, None, None, None, None, None, None, None, None, 
			None, None, None, None, None, None, None], ws, paint=False))
		
		ws.merge_cells(f'A{rows_count+3}:D{rows_count+3}')
		bd = Side(border_style="thin", color="ffffff")
		ws[f"A{rows_count+3}"].border = Border(left=bd, top=bd, right=bd, bottom=bd)
		ws[f"B{rows_count+3}"].border = Border(left=bd, top=bd, right=bd, bottom=bd)
		ws[f"C{rows_count+3}"].border = Border(left=bd, top=bd, right=bd, bottom=bd)
		ws[f"D{rows_count+3}"].border = Border(left=bd, top=bd, right=bd, bottom=bd)

		rows_count += 3

		if qualis_year == "2016":
			indicators = self.authors_indicators_2016_proceedings
		else:
			indicators = self.authors_indicators_2019_proceedings
		ws, rows_count = self.build_separated_summary(ws, authors_list, qualis_year, rows_count, indicators, pub_type="proceedings")
		

	def add_artprof_sheet(self):
		ws = self.create_sheet("Art|Prof")
		for row in dataframe_to_rows(self.art_prof, index=False, header=True):
			ws.append(row)

	def add_artppg_sheet(self):
		ws = self.create_sheet("Art|PPG")
		for row in dataframe_to_rows(self.artppg, index=False, header=True):
			ws.append(row)
		ws.append([""])
		for row in dataframe_to_rows(self.general_indicators_2016, index=False, header=True):
			ws.append(self.styled_cells(row, ws, qualis_year="2016"))
		ws.append([""])
		for row in dataframe_to_rows(self.general_indicators_2019, index=False, header=True):
			ws.append(self.styled_cells(row, ws, qualis_year="2019"))

		ws.append([None, None, None])
		ws.append([None, None, None])
		ws.append([None, None, None])

	def add_proceedingsppg_sheet(self):
		ws = self.create_sheet("Anais|PPG")
		df = pd.DataFrame()
		df["Nome de Publicação"] = self.proceedings["Nome de Publicação"]
		df["Sigla"] = self.proceedings["SIGLA"]
		df["Qualis CC 2016"] = self.proceedings["Qualis CC 2016"]
		df["Qualis 2019"] = self.proceedings["Qualis 2019"]
		df["Quantidade"] = self.proceedings["Quantidade"]
		sum_ = 0
		for i in df["Quantidade"]:
			sum_ += i
		percentages = []
		for i in df["Quantidade"]:
			percentages.append(f"{round(100/sum_ * i, 1)}%")
		df["Porcentagem"] = percentages

		for row in dataframe_to_rows(df, index=False, header=True):
			ws.append(row)
		ws.append([None])

		for row in dataframe_to_rows(self.proceedings_metrics_2016, index=False, header=True):
			ws.append(self.styled_cells(row, ws, paint=False))
		ws.append([None])
		for row in dataframe_to_rows(self.proceedings_metrics_2019, index=False, header=True):
			ws.append(self.styled_cells(row, ws, paint=False))
			
		graphs = Graphs_Proceedings_Journals(df.copy(), "Anais de Eventos Utilizados para Publicação")
		ws = graphs.add_graphs(ws)

	def add_journalsppg_sheet(self):
		ws = self.create_sheet("Periódicos|PPG")
		df = pd.DataFrame()
		df["Nome de Publicação"] = self.journals["Nome de Publicação"]
		df["ISSN"] = self.journals["ISSN"]
		df["Qualis CC 2016"] = self.journals["Qualis CC 2016"]
		df["Qualis 2019"] = self.journals["Qualis 2019"]
		df["Scopus 2019"] = self.journals["Scopus 2019"]
		df["Quantidade"] = self.journals["Quantidade"]
		sum_ = 0
		for i in df["Quantidade"]:
			sum_ += i
		percentages = []
		for i in df["Quantidade"]:
			percentages.append(f"{round(100/sum_ * i, 1)}%")
		df["Porcentagem"] = percentages

		for row in dataframe_to_rows(df, index=False, header=True):
			ws.append(row)
		ws.append([None])
		for row in dataframe_to_rows(self.journal_metrics_2016, index=False, header=True):
			ws.append(self.styled_cells(row, ws, paint=False))
		ws.append([None])
		for row in dataframe_to_rows(self.journal_metrics_2019, index=False, header=True):
			ws.append(self.styled_cells(row, ws, paint=False))


		graphs = Graphs_Proceedings_Journals(df.copy(), "Periódicos Utilizados para Publicação")
		ws = graphs.add_graphs(ws)

	def add_authors_sheets(self):
		for pos, author in enumerate(self.reports["Author"]):
			row_count = 1
			if FILE == "EGRESSOS 2017-2020":
				first_name = author.split(" ")[0]
				second_name = author.split(" ")[1]
				if len(second_name) < 3:
					second_name = second_name + " " + author.split(" ")[2]
				author_name = first_name + " " + second_name

				ws = self.create_sheet(author_name)
			else:
				ws = self.create_sheet(author.split(" ")[0])

			for row in dataframe_to_rows(self.reports["Report"][pos], index=False, header=True): # Adiciona o dataframe ao sheet
				ws.append(row)
				row_count += 1

			ws.append([None, None, None])
			ws.append([None, None, None])
			ws.append(self.styled_cells(["Periódicos e Anais - Qualis 2016"], ws, paint=False))
			row_count += 2
			ws.merge_cells(f'A{row_count}:B{row_count}')
			ws.append([None, None, None])
			row_count += 1
			for row in dataframe_to_rows(self.authors_indicators_2016[pos], index=False, header=True):
				ws.append(self.styled_cells(row, ws, qualis_year="2016"))
				row_count += 1

			ws.append([None, None, None])
			ws.append([None, None, None])
			ws.append(self.styled_cells(["Periódicos e Anais - Qualis 2019"], ws, paint=False))
			row_count += 3
			ws.merge_cells(f'A{row_count}:B{row_count}')
			ws.append([None, None, None])
			for row in dataframe_to_rows(self.authors_indicators_2019[pos], index=False, header=True):
				ws.append(self.styled_cells(row, ws))

			ws.append([None, None, None])
			average = Cell(ws, column="A", row=1, value=self.averages[pos])
			average.font = Font(bold=True)
			ws.append([None, average])
			#ws.append([self.averages[pos]])
			ws.append([None, None, None])
			ws.append([None, None, None])
			ws.append([None, None, None])

	def add_exceptions_sheet(self):
		ws = self.create_sheet("Exceções")
		pos_artigos = False
		for pos, row in enumerate(dataframe_to_rows(self.exceptions, index=False, header=True)):
			if "Nome Periódico" in row and "Qualis 2019/ISSN Impresso" in row and "Qualis 2019/ISSN Online" in row:
				ws.append(self.styled_cells(row, ws, paint=False))
				pos_artigos = True
			elif pos_artigos == True:
				ws.append(self.styled_cells(row, ws, paint=False))
			else:
				ws.append(row)

	def add_info(self):
		self.add_parameters_sheet()
		self.add_indicators_sheet()
		self.add_authors_sheet()
		self.add_artppg_sheet()
		authors_list = self.get_authors_list()	
		self.add_graphs_sheet(authors_list)
		self.add_summary_sheet(authors_list, "2016")
		self.add_summary_sheet(authors_list, "2019")
		self.add_artprof_sheet()
		if HAS_EVENTS == True:
			self.add_proceedingsppg_sheet()
		self.add_journalsppg_sheet()
		self.add_authors_sheets()
		if HAS_EVENTS == True:
			self.add_exceptions_sheet()


	# def altera_authors(self):
	# 	authors_list = []
	# 	egress_list = []
	# 	students_list = []

	# 	for autor in self.authors["Autor"]:
	# 		authors_list.append(autor)
	# 	for egress in self.egress_list:
	# 		egress_list.append(egress.name)
	# 	for student in self.students_list:
	# 		students_list.append(student.name)

	# 	for ws in self.worksheets:
	# 		for col in ws.columns:
	# 			if (ws.title == "Art|Prof" and col[0].column_letter not in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']) or (ws.title != "Art|Prof" and col[0].column_letter not in ['A', 'B', 'C', 'D', 'E', 'F', 'G']):
	# 				for pos, cell in enumerate(col):
	# 					if cell.row != 1 and str(cell.value) != "" and str(cell.value) != " " and cell.value != None and str(cell.value) != 'nan':
	# 						temp, cell.value = search_authors_list(authors_list, str(cell.value))
	# 						temp, cell.value = search_authors_list(egress_list, str(cell.value))
	# 						temp, cell.value = search_authors_list(students_list, str(cell.value))


	def apply_style(self):
		for ws in self.worksheets:
			# Add a border to the cells
			for col in ws.columns:
				for cell in col:
					if cell.value != None and str(cell.value) != "nan" and "Média de autores" not in str(cell.value) and cell.value != "Número médio de docentes" and cell.value != "Periódicos e Anais - Qualis 2016" and cell.value != "Periódicos e Anais - Qualis 2019" and cell.value != "Periódicos - Qualis 2016" and cell.value != "Periódicos - Qualis 2019" and cell.value != "Anais - Qualis 2016" and cell.value != "Anais - Qualis 2019" and cell.value != "Qualis 2013-2016" and cell.value != "Qualis 2017-2020":
						bd = Side(border_style="thin", color="000000")
						cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)

			# Style of the column titles
			for cell in ws[1]:
				cell.font = Font(bold=True) # Bold
				if ws.title != "Indicadores":
					cell.alignment = Alignment(horizontal='center', vertical='center') # Center 

			# Center columns
			for col in ws.columns:
				if ws.title == "Resumo Q2016" or ws.title == "Resumo Q2019":
					for cell in col:
						cell.alignment = Alignment(horizontal='center', vertical='center')

				elif ws.title == "Anais|PPG" or ws.title =="Periódicos|PPG" or ws.title == "Autores": # For these sheets all the columns are centered except for the first one
					if col[0].column_letter != "A":
						for cell in col:
							cell.alignment = Alignment(horizontal='center', vertical='center')

				elif ws.title == "Art|Prof":
					if col[0].column_letter not in 'DE': # If its not the "Título" or "Nome de Publicação" columns
						for cell in col:
							cell.alignment = Alignment(horizontal='center', vertical='center')

				elif ws.title == "Indicadores":
					if col[0].column_letter not in 'AE': 
						for cell in col:
							cell.alignment = Alignment(horizontal='center', vertical='center')

				elif ws.title != "Exceções": # For all the other sheets except for "Exceções"
					if col[0].column_letter not in 'CD': # If its not the "Título" or "Nome de Publicação" columns
						for cell in col:
							if cell.value != "Número médio de docentes":
								cell.alignment = Alignment(horizontal='center', vertical='center')
					
			if ws.title == "Indicadores":
				bold_cells = ["A1", "A2", "B2", "C2", "E2", "F2", "G2", "A9", "A10", "B10", "C10", "E10", "F10", "G10"]
				for cell in bold_cells:
					ws[cell].font = Font(bold=True)

			if ws.title == "Resumo Q2016" or ws.title == "Resumo Q2019":
				for row in ws.rows:
					bold = False
					for pos, cell in enumerate(row):
						if pos == 0:
							if cell.value == 'PPGtotal' or cell.value == 'PPGdoc':
								bold = True
						if bold == True:
							cell.font = Font(bold=True)


				col_titles = ['Índices', 'Publicações totais', 'Publicações com alunos/egressos', 'Nome', 'Autores/Artigo', 'Irestrito', 'Igeral', 'Periódicos', 'Anais', 'A1-B1', 'A1', 'A2', 'B1', 'B2-B5', 'B2', 'B3', 'B4', 'B5', 'Outros', 'Periódicos', 'Anais', 'A1-B1', 'A1', 'A2', 'B1', 'B2-B5', 'B2', 'B3', 'B4', 'B5', 'Outros', 'Periódicos A1-B1', 'Periódicos A1-B1 com alunos/egressos']
				for col in ws.columns:
					for pos, cell in enumerate(col):

						if "A/E" in str(cell.value):
							cell.value = str(cell.value).replace(" A/E", "")

						if cell.value in col_titles:
							cell.font = Font(bold=True)
						
						if pos > 1 and pos <= self.summary_size:
							if "2016" in ws.title:
								if cell.column in [3, 5, 7, 8, 9, 10, 17, 19, 20, 21, 22, 29, 30]:
									cell.font = Font(color='FFFFFF', bold=True)
							else:
								if cell.column in [3, 5, 7, 8, 9, 10, 11, 18, 20, 21, 22, 23, 24, 31, 32]:
									cell.font = Font(color='FFFFFF', bold=True)

						if pos >= self.summary_size+4: # because of the blank rows

							if pos != self.summary_size*2 + 5:
								if "2016" in ws.title:
									if cell.column in [3, 5, 6, 7, 8, 15, 16, 17, 18]:
										cell.font = Font(color='FFFFFF', bold=True)
								else:
									if cell.column in [3, 5, 6, 7, 8, 9, 16, 17, 18, 19, 20]:
										cell.font = Font(color='FFFFFF', bold=True)



	def search_students(self, col, ws):
		for row, cell in enumerate(col):
			for egress in self.egress_list:
				if str(cell.value).lower() == egress.name.lower():
					if ws.title == "Art|Prof":
						if egress.period[str(int(ws.cell(row=row+1, column=2).value))[2:4]] == True and int(str(int(ws.cell(row=row+1, column=2).value))[2:]) > egress.egress_year:
							cell.fill = PatternFill(fill_type='solid', start_color='00FFFF', end_color='00FFFF')
							ws.cell(row=row+1, column=10).value = "X"
						elif egress.period[str(int(ws.cell(row=row+1, column=2).value))[2:4]] == True:
							cell.fill = PatternFill(fill_type='solid', start_color='F781F3', end_color='F781F3')
							ws.cell(row=row+1, column=10).value = "X"
					elif ws.title == "Art|PPG":
						if egress.period[str(int(ws.cell(row=row+1, column=1).value))[2:4]] == True and int(str(int(ws.cell(row=row+1, column=1).value))[2:]) > egress.egress_year:
							cell.fill = PatternFill(fill_type='solid', start_color='00FFFF', end_color='00FFFF')
							ws.cell(row=row+1, column=10).value = "X"
						elif egress.period[str(int(ws.cell(row=row+1, column=1).value))[2:4]] == True:
							cell.fill = PatternFill(fill_type='solid', start_color='F781F3', end_color='F781F3')
							ws.cell(row=row+1, column=10).value = "X"
					else:
						if egress.period[str(int(ws.cell(row=row+1, column=1).value))[2:4]] == True and int(str(int(ws.cell(row=row+1, column=1).value))[2:]) > egress.egress_year:
							cell.fill = PatternFill(fill_type='solid', start_color='00FFFF', end_color='00FFFF')
							ws.cell(row=row+1, column=9).value = "X"
						elif egress.period[str(int(ws.cell(row=row+1, column=1).value))[2:4]] == True:
							cell.fill = PatternFill(fill_type='solid', start_color='F781F3', end_color='F781F3')
							ws.cell(row=row+1, column=9).value = "X"
			for student in self.students_list:
				if str(cell.value).lower() == student.name.lower():
					if ws.title == "Art|Prof":
						if student.period[str(ws.cell(row=row+1, column=2).value)[2:4]] == True:
							cell.fill = PatternFill(fill_type='solid', start_color='F781F3', end_color='F781F3')
							ws.cell(row=row+1, column=10).value = "X"
					elif ws.title == "Art|PPG":
						if student.period[str(ws.cell(row=row+1, column=1).value)[2:4]] == True:
							cell.fill = PatternFill(fill_type='solid', start_color='F781F3', end_color='F781F3')
							ws.cell(row=row+1, column=10).value = "X"
					else:
						if student.period[str(ws.cell(row=row+1, column=1).value)[2:4]] == True:
							cell.fill = PatternFill(fill_type='solid', start_color='F781F3', end_color='F781F3')
							ws.cell(row=row+1, column=9).value = "X"


	def apply_colors(self):
		for ws in self.worksheets:

			if ws.title == "Anais|PPG" or ws.title == "Periódicos|PPG":
				for col in ws.columns: 
					if col[0].column_letter in "CDE":
						for cell in col:
							if str(cell.value) in ["A1", "A2", "A3", "A4"]:
								cell.font = Font(color='FFFAFA')
								cell.fill = PatternFill(fill_type='solid', start_color='008000', end_color='008000')
							elif str(cell.value) in ["B1", "B2", "B3", "B4"]:
								cell.fill = PatternFill(fill_type='solid', start_color='FFFF99', end_color='FFFF99')
							elif "Q1" in str(cell.value) or "Q2" in str(cell.value):
								cell.font = Font(color='FFFAFA')
								cell.fill = PatternFill(fill_type='solid', start_color='008000', end_color='008000')
							elif "Q3" in str(cell.value) or "Q4" in str(cell.value):
								cell.fill = PatternFill(fill_type='solid', start_color='FFFF99', end_color='FFFF99')	

			elif ws.title == "Gráficos Q2016" or ws.title == "Gráficos Q2019":
				for i in range(7):
					for j in range(50):
						cell = ws.cell(row=j+1, column=i+1, value=None)
						cell.fill = PatternFill(fill_type='solid', start_color='FFFFFF', end_color='FFFFFF')

			elif ws.title == "Resumo Q2016" or ws.title == "Resumo Q2019":
				if "2016" in ws.title:
					green_general = ["C", "E", "G", "H", "I", "J", "Q", "S", "T", "U", "V", "AC", "AD"]
					white_general = ["A", "B", "P", "AB"]
					green_separated = ["C", "E", "F", "G", "H", "O", "P", "Q", "R"]
					white_separated = ["A", "B", "N", "X"]
				else:
					green_general = ["C", "E", "G", "H", "I", "J", "K", "R", "T", "U", "V", "W", "X", "AE", "AF"]
					white_general = ["A", "B", "Q", "AD"]
					green_separated = ["C", "E", "F", "G", "H", "I", "P", "Q", "R", "S", "T"]
					white_separated = ["A", "B", "O", "Z"]
				for col in ws.columns:
					for pos, cell in enumerate(col):
						if pos > 1 and pos <= self.summary_size: # General Summary
							if col[2].column_letter in green_general:
								cell.fill = PatternFill(fill_type='solid', start_color='00B050', end_color='008000')
							elif col[2].column_letter not in white_general:
								cell.fill = PatternFill(fill_type='solid', start_color='FFFFCC', end_color='FFFF99')

						elif pos >= self.summary_size+4: # because of the blank lines
							if pos != self.summary_size*2 + 5:
								if col[2].column_letter in green_separated:
									cell.fill = PatternFill(fill_type='solid', start_color='00B050', end_color='008000')
								elif col[2].column_letter not in white_separated:
									cell.fill = PatternFill(fill_type='solid', start_color='FFFFCC', end_color='FFFF99')

						if cell.value is None or str(cell.value) =="nan":
							cell.fill = PatternFill(fill_type='solid', start_color='FFFFFF', end_color='FFFFFF')


			else:
				if ws.title == "Art|Prof":
					type_column = "C"
					Q16_column = "G"
					Q20_column = "H"
					scopus_column = "I"
					# Q16_column = "GH"
					# Q20_column = "I"
				else:
					type_column = "B"
					Q16_column = "F"
					Q20_column = "G"
					scopus_column = "H"
					# Q16_column = "FG"
					# Q20_column = "H"

				
				# Fill all the cells with white
				for col in ws.iter_cols(min_col=None, max_col=None, min_row=None, max_row=None):
					for cell in col:
						if cell.value is None or str(cell.value) =="nan":
							cell.fill = PatternFill(fill_type='solid', start_color='FFFFFF', end_color='FFFFFF')
				

				for col in ws.columns:
					# Searches for students and egress
					if col[0].column_letter not in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
						self.search_students(col, ws)

					if col[0].column_letter in type_column:
						for cell in col:
							if cell.value == "Periódico":
								cell.font = Font(color='FFFAFA')
								cell.fill = PatternFill(fill_type='solid', start_color='008000', end_color='008000')
							elif cell.value == "Anais":
								cell.fill = PatternFill(fill_type='solid', start_color='FFFF99', end_color='FFFF99')
							elif cell.value == "Livros":
								cell.fill = PatternFill(fill_type='solid', start_color='FF8000', end_color='FF8000')
							elif cell.value == "Capítulos":
								cell.fill = PatternFill(fill_type='solid', start_color='F7D358', end_color='F7D358')

					elif col[0].column_letter in Q16_column:
						for cell in col:
							if "A1" in str(cell.value) or "A2" in str(cell.value) or "B1" in str(cell.value):
								cell.font = Font(color='FFFAFA')
								cell.fill = PatternFill(fill_type='solid', start_color='008000', end_color='008000')
							elif "B2" in str(cell.value) or "B3" in str(cell.value) or "B4" in str(cell.value) or "B5" in str(cell.value):
								cell.fill = PatternFill(fill_type='solid', start_color='FFFF99', end_color='FFFF99')

					elif col[0].column_letter in Q20_column:
						for cell in col:
							if "A1" in str(cell.value) or "A2" in str(cell.value) or "A3" in str(cell.value) or "A4" in str(cell.value):
								cell.font = Font(color='FFFAFA')
								cell.fill = PatternFill(fill_type='solid', start_color='008000', end_color='008000')
							elif "B1" in str(cell.value) or "B2" in str(cell.value) or "B3" in str(cell.value) or "B4" in str(cell.value):
								cell.fill = PatternFill(fill_type='solid', start_color='FFFF99', end_color='FFFF99')

					elif col[0].column_letter in scopus_column:
						for cell in col:
							if "Q1" in str(cell.value) or "Q2" in str(cell.value):
								cell.font = Font(color='FFFAFA')
								cell.fill = PatternFill(fill_type='solid', start_color='008000', end_color='008000')
							elif "Q3" in str(cell.value) or "Q4" in str(cell.value):
								cell.fill = PatternFill(fill_type='solid', start_color='FFFF99', end_color='FFFF99')

	def apply_filters(self):
		for ws in self.worksheets:
			if ws.title != "Parâmetros" and ws.title != "Indicadores" and ws.title != "Art|Prof" and ws.title != "Gráficos Q2016" and ws.title != "Gráficos Q2019" and ws.title != "Resumo Q2016" and ws.title != "Resumo Q2019" and ws.title != "Anais|PPG" and ws.title != "Periódicos|PPG" and ws.title != "Exceções":
				rows = ws.dimensions[4:]
				rows = int(rows) -56
				dimension = ws.dimensions[:4] + str(rows)
				ws.auto_filter.ref = dimension
			elif ws.title == "Anais|PPG" or ws.title == "Periódicos|PPG":
				rows = ws.dimensions[4:]
				rows = int(rows) -12
				dimension = ws.dimensions[:4] + str(rows)
				ws.auto_filter.ref = dimension

			elif ws.title != "Gráficos Q2016" and ws.title != "Gráficos Q2019" and ws.title != "Resumo Q2016" and ws.title != "Resumo Q2019" and ws.title != "Indicadores" and ws.title != "Parâmetros":
				ws.auto_filter.ref = ws.dimensions


	def apply_dimensions(self):
		for ws in self.worksheets:
			if ws.title == "Exceções":
				ws.column_dimensions['A'].width = 50
				ws.column_dimensions['B'].width = 50
				size = len(self.exceptions["Nome Evento Canônico"][0])
				for i in self.exceptions["Nome Evento Canônico"]:
					if len(str(i)) > size:
						size = len(i)
				ws.column_dimensions['C'].width = size

			elif ws.title == "Parâmetros":
				ws.column_dimensions['A'].width = 28
				ws.column_dimensions['B'].width = 15

			elif ws.title == "Indicadores":
				ws.column_dimensions['A'].width = 20
				ws.column_dimensions['B'].width = 15
				ws.column_dimensions['C'].width = 15
				ws.column_dimensions['D'].width = 10
				ws.column_dimensions['E'].width = 20
				ws.column_dimensions['F'].width = 15
				ws.column_dimensions['G'].width = 15

			elif ws.title == "Gráficos Q2016" or ws.title == "Gráficos Q2019":
				ws.column_dimensions['A'].width = 30
				ws.column_dimensions['B'].width = 30
				ws.column_dimensions['C'].width = 30
				ws.column_dimensions['D'].width = 30
				ws.column_dimensions['E'].width = 30
				ws.column_dimensions['F'].width = 30
				ws.column_dimensions['G'].width = 30
				for i in range(50):
					ws.row_dimensions[i+1].height = 50

			elif ws.title == "Anais|PPG":
				ws.column_dimensions['A'].width = 60
				ws.column_dimensions['B'].width = 15
				ws.column_dimensions['C'].width = 18
				ws.column_dimensions['D'].width = 18
				ws.column_dimensions['E'].width = 18
				ws.column_dimensions['F'].width = 15
				ws.column_dimensions['G'].width = 15

			elif ws.title == "Periódicos|PPG":
				ws.column_dimensions['A'].width = 60
				ws.column_dimensions['B'].width = 15
				ws.column_dimensions['C'].width = 18
				ws.column_dimensions['D'].width = 18
				ws.column_dimensions['E'].width = 18
				ws.column_dimensions['F'].width = 18
				ws.column_dimensions['G'].width = 15

			elif ws.title == "Resumo Q2016" or ws.title == "Resumo Q2019":
				if FILE == "EGRESSOS 2017-2020":
					ws.column_dimensions['A'].width = 25
				else:
					ws.column_dimensions['A'].width = 15
				ws.column_dimensions['B'].width = 15
				if "2016" in ws.title:
					columns = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", 
					"N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", 
					"AA", "AB"]
				else:
					columns = ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", 
					"N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", 
					"AA", "AB", "AC", "AD"]
				for column in columns:
					ws.column_dimensions[column].width = 10

				if "2016" in ws.title:
					ws.column_dimensions["AC"].width = 35
					ws.column_dimensions["AD"].width = 35
				else:
					ws.column_dimensions["AE"].width = 35
					ws.column_dimensions["AF"].width = 35

			elif ws.title == "Autores":
				for column in ws.columns:
					if column[0].column_letter == "A":
						ws.column_dimensions['A'].width = 50
					else:
						ws.column_dimensions[column[0].column_letter].width = 10
					
				ws.row_dimensions[1].height = 25 # Altura da linha dos nomes das colunas


			else:
				dic = {1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J', 11: 'K', 12: 'L', 13 : 'M', 14: 'N', 15: 'O', 16: 'P', 17: 'Q', 18: 'R', 19: 'S', 20: 'T', 21: 'U', 22: 'V', 23: 'W', 24: 'X', 25: 'Y'}

				if ws.title == "Art|PPG":
					ws.column_dimensions['A'].width = 20
					ws.column_dimensions['B'].width = 15
					ws.column_dimensions['C'].width = 35
					ws.column_dimensions['D'].width = 35
					ws.column_dimensions['E'].width = 20
					ws.column_dimensions['F'].width = 18
					ws.column_dimensions['G'].width = 18
					ws.column_dimensions['H'].width = 18
					ws.column_dimensions['I'].width = 18
					ws.column_dimensions['J'].width = 10
					ws.column_dimensions['K'].width = 30

					i = 12
					while i <= ws.max_column:
						ws.column_dimensions[dic[i]].width = 30
						i += 1

				elif ws.title == "Art|Prof":
					ws.column_dimensions['A'].width = 35
					ws.column_dimensions['B'].width = 15
					ws.column_dimensions['C'].width = 15
					ws.column_dimensions['D'].width = 35
					ws.column_dimensions['E'].width = 35
					ws.column_dimensions['F'].width = 15
					ws.column_dimensions['G'].width = 18
					ws.column_dimensions['H'].width = 18
					ws.column_dimensions['I'].width = 18
					ws.column_dimensions['J'].width = 10
					ws.column_dimensions['K'].width = 30

					i = 12
					while i <= ws.max_column:
						ws.column_dimensions[dic[i]].width = 30
						i += 1
			
					ws.row_dimensions[1].height = 25 # Altura da linha dos nomes das colunas

				else:
					ws.column_dimensions['A'].width = 20
					ws.column_dimensions['B'].width = 15
					ws.column_dimensions['C'].width = 35
					ws.column_dimensions['D'].width = 35
					ws.column_dimensions['E'].width = 20
					ws.column_dimensions['F'].width = 18
					ws.column_dimensions['G'].width = 18
					ws.column_dimensions['H'].width = 18
					ws.column_dimensions['I'].width = 10
					ws.column_dimensions['J'].width = 30

					i = 11
					while i <= ws.max_column:
						ws.column_dimensions[dic[i]].width = 30
						i += 1
			
					ws.row_dimensions[1].height = 25 # Altura da linha dos nomes das colunas
			

	def save_file(self):
		self.save(EXCEL_FILE_NAME) # Salva o arquivo
		for file in os.listdir("temp"):
			os.remove(f"temp/{file}") # Remove the file
		os.rmdir("temp") # Remove the temporary folder

		for file in os.listdir("temp2"):
			os.remove(f"temp2/{file}") # Remove the file
		os.rmdir("temp2") # Remove the temporary folder
		